from datetime import datetime, timedelta
import threading
import app as _app
from app.models import Execution, ExecutionStatus, ParticipantStatus, WorkflowStep
from app.services.email_service import EmailService
from app.services.token_service import TokenService
import logging


def _db():
    """Accesso dinamico a db_session (evita riferimento stale)"""
    return _app.db_session

logger = logging.getLogger(__name__)


class SchedulerService:
    """Enhanced Scheduler Service with wait_until support"""
    
    @staticmethod
    def schedule_step(participant, step, delay_hours=0):
        """
        Schedule execution of a step
        
        Supports:
        - delay_hours: Wait X hours before execution
        - wait_until: Wait until specific date/time (for wait_until step type)
        
        Args:
            participant: Participant instance
            step: WorkflowStep instance
            delay_hours: Hours to delay (default 0 = immediate)
            
        Returns:
            Execution instance
        """
        try:
            # Calculate scheduled time
            scheduled_at = None
            
            # Check if this is a wait_until step
            if step.type.value == 'wait_until':
                scheduled_at = SchedulerService._calculate_wait_until(step)
                logger.info(f"Wait Until step calculated: {scheduled_at}")
            else:
                # Regular delay
                scheduled_at = datetime.utcnow() + timedelta(hours=delay_hours)
            
            # Create execution record
            execution = Execution(
                participant_id=participant.id,
                step_id=step.id,
                status=ExecutionStatus.SCHEDULED,
                scheduled_at=scheduled_at
            )
            
            _db().add(execution)
            _db().flush()  # Get ID

            # Unique job ID
            job_id = f"exec_{execution.id}_{participant.id}_{step.id}"
            execution.job_id = job_id

            # Commit prima di schedulare il job (il job gira in un thread separato
            # e ha bisogno di trovare l'Execution nel DB)
            _db().commit()

            # Schedule job
            now = datetime.utcnow()
            if scheduled_at <= now:
                # Esecuzione immediata in thread separato (evita problemi timezone con APScheduler)
                def _run_delayed(exec_id):
                    import time
                    time.sleep(3)  # Attende che il chiamante faccia commit
                    SchedulerService._execute_step(exec_id)

                t = threading.Thread(target=_run_delayed, args=[execution.id], daemon=True)
                t.start()
                logger.info(f"✓ Scheduled step {step.id} for immediate execution (3s delay)")
            else:
                # Converti UTC → locale per APScheduler (configurato Europe/Rome)
                import pytz
                local_tz = pytz.timezone('Europe/Rome')
                run_date_local = pytz.utc.localize(scheduled_at).astimezone(local_tz)

                _app.scheduler.add_job(
                    func=SchedulerService._execute_step,
                    args=[execution.id],
                    trigger='date',
                    run_date=run_date_local,
                    id=job_id,
                    replace_existing=True
                )
                time_until = (scheduled_at - now).total_seconds() / 3600
                logger.info(f"✓ Scheduled step {step.id} for {run_date_local.strftime('%d/%m/%Y %H:%M')} local ({time_until:.1f}h from now)")

            return execution

        except Exception as e:
            _db().rollback()
            logger.error(f"✗ Error scheduling step: {str(e)}")
            raise
    
    @staticmethod
    def _calculate_wait_until(step):
        """
        Calculate target datetime for wait_until step.
        L'utente inserisce orari in ora locale (Europe/Rome),
        vengono convertiti in UTC per lo storage.

        Returns:
            datetime: Target execution time (UTC)
        """
        import pytz

        config = step.skip_conditions or {}
        wait_type = config.get('wait_type', 'date')

        local_tz = pytz.timezone('Europe/Rome')
        now_local = datetime.now(local_tz)
        now_utc = datetime.utcnow()

        if wait_type == 'date':
            target_date = config.get('target_date')  # YYYY-MM-DD
            target_time = config.get('target_time', '09:00')  # HH:MM

            if not target_date:
                logger.warning("No target_date specified for wait_until, using current time")
                return now_utc

            try:
                datetime_str = f"{target_date} {target_time}"
                naive_dt = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
                # Interpreta come ora locale e converti in UTC
                local_dt = local_tz.localize(naive_dt)
                utc_dt = local_dt.astimezone(pytz.utc).replace(tzinfo=None)

                if utc_dt < now_utc:
                    logger.warning(f"Target date {naive_dt} (local) is in the past, executing immediately")
                    return now_utc

                return utc_dt

            except ValueError as e:
                logger.error(f"Error parsing wait_until date: {e}")
                return now_utc

        elif wait_type == 'time':
            target_time = config.get('target_time', '09:00')  # HH:MM

            try:
                hour, minute = map(int, target_time.split(':'))
                target_local = now_local.replace(hour=hour, minute=minute, second=0, microsecond=0)

                if target_local <= now_local:
                    target_local += timedelta(days=1)

                utc_dt = target_local.astimezone(pytz.utc).replace(tzinfo=None)
                return utc_dt

            except (ValueError, AttributeError) as e:
                logger.error(f"Error parsing time: {e}")
                return now_utc

        elif wait_type == 'day_of_week':
            target_day = config.get('target_day', 'monday').lower()
            target_time = config.get('target_time', '09:00')

            days_map = {
                'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
                'friday': 4, 'saturday': 5, 'sunday': 6
            }

            target_weekday = days_map.get(target_day, 0)
            current_weekday = now_local.weekday()

            days_ahead = target_weekday - current_weekday
            if days_ahead <= 0:
                days_ahead += 7

            target_local = now_local + timedelta(days=days_ahead)

            try:
                hour, minute = map(int, target_time.split(':'))
                target_local = target_local.replace(hour=hour, minute=minute, second=0, microsecond=0)
                utc_dt = target_local.astimezone(pytz.utc).replace(tzinfo=None)
                return utc_dt
            except (ValueError, AttributeError) as e:
                logger.error(f"Error parsing time for day_of_week: {e}")
                return now_utc

        elif wait_type == 'delay_hours':
            delay = config.get('delay_hours', 0)
            return now_utc + timedelta(hours=int(delay))

        else:
            logger.warning(f"Unknown wait_type: {wait_type}, executing immediately")
            return now_utc
    
    @staticmethod
    def _execute_step(execution_id):
        """
        Execute a step (called by scheduler)

        Args:
            execution_id: Execution ID
        """
        from app import create_app

        # Need app context for DB operations (riusa l'app esistente, non ricrea)
        app = create_app()
        
        with app.app_context():
            try:
                execution = _db().get(Execution, execution_id)
                if not execution:
                    logger.error(f"✗ Execution {execution_id} not found")
                    return
                
                participant = execution.participant
                step = execution.step
                
                logger.info(f"▶ Executing step {step.id} for participant {participant.id}")
                
                # Check if participant already completed
                if participant.status == ParticipantStatus.COMPLETED:
                    execution.status = ExecutionStatus.SKIPPED
                    execution.result_data = {'reason': 'Participant already completed'}
                    _db().commit()
                    logger.info(f"⊘ Skipped: participant {participant.id} already completed")
                    return
                
                # Check skip conditions
                if SchedulerService._should_skip(participant, step):
                    execution.status = ExecutionStatus.SKIPPED
                    execution.result_data = {'reason': 'Skip conditions met'}
                    _db().commit()
                    logger.info(f"⊘ Skipped: conditions met for step {step.id}")
                    return
                
                # Execute based on type
                success = False
                
                if step.type.value == 'email':
                    success = SchedulerService._execute_email_step(participant, step, execution)
                    # If wait_for_landing is enabled, handle waiting internally
                    email_config = step.skip_conditions or {}
                    if success and email_config.get('wait_for_landing'):
                        execution.status = ExecutionStatus.SENT
                        execution.sent_at = datetime.utcnow()
                        participant.last_interaction = datetime.utcnow()
                        _db().commit()
                        SchedulerService._schedule_landing_wait(participant, step, email_config)
                        return  # Don't use default next-step logic
                elif step.type.value == 'wait_until':
                    # Wait until step just triggers next step
                    success = True
                    logger.info(f"✓ Wait Until step completed, triggering next step")
                elif step.type.value == 'goal_check':
                    # Goal check handles its own branching (jump/complete/skip)
                    success = SchedulerService._execute_goal_check_step(participant, step, execution)
                    gc_action = (step.skip_conditions or {}).get(
                        'if_met' if execution.result_data and execution.result_data.get('goal_met') else 'if_not_met',
                        'continue'
                    )
                    if gc_action in ('jump', 'complete', 'skip'):
                        # Already handled scheduling internally
                        if not success:
                            execution.status = ExecutionStatus.FAILED
                            execution.error_message = "Goal check failed"
                            _db().commit()
                        return
                elif step.type.value == 'condition':
                    # Condition step — branching
                    success = SchedulerService._execute_condition_step(participant, step, execution)
                    # Condition handles its own next-step scheduling
                    if success:
                        execution.status = ExecutionStatus.SENT
                        execution.sent_at = datetime.utcnow()
                        participant.last_interaction = datetime.utcnow()
                    else:
                        execution.status = ExecutionStatus.FAILED
                        execution.error_message = "Condition evaluation failed"
                    _db().commit()
                    return  # Don't use default next-step logic
                elif step.type.value == 'human_approval':
                    success = SchedulerService._execute_human_approval_step(participant, step, execution)
                elif step.type.value == 'survey':
                    success = SchedulerService._execute_survey_step(participant, step, execution)
                elif step.type.value == 'export_data':
                    # Export data step
                    success = SchedulerService._execute_export_data_step(participant, step, execution)
                elif step.type.value == 'whatsapp':
                    success = SchedulerService._execute_whatsapp_step(participant, step, execution)
                elif step.type.value == 'excel_write':
                    success = SchedulerService._execute_excel_write_step(participant, step, execution)
                else:
                    logger.warning(f"⚠ Unsupported step type: {step.type}")

                # Update status
                if success:
                    execution.status = ExecutionStatus.SENT
                    execution.sent_at = datetime.utcnow()
                    participant.last_interaction = datetime.utcnow()

                    # Schedule next step
                    SchedulerService._schedule_next_step(participant, step)
                else:
                    execution.status = ExecutionStatus.FAILED
                    execution.error_message = execution.error_message or f"{step.type.value} step failed"
                
                _db().commit()
                
            except Exception as e:
                logger.error(f"✗ Error executing step {execution_id}: {str(e)}")
                _db().rollback()
    
    @staticmethod
    def _schedule_landing_wait(participant, step, config):
        """
        Mark participant as waiting for landing page submission.
        Actual polling is handled by the global cron job check_all_landing_waits().
        """
        logger.info(f"✓ Landing wait registered for participant {participant.id}, "
                     f"timeout in {config.get('landing_timeout_days', 7)} days")

    @staticmethod
    def check_all_landing_waits():
        """
        Global cron job: checks ALL participants waiting for landing page submission.
        Runs every 10 minutes via APScheduler. Survives server restarts.
        """
        from app.models import Participant
        try:
            # Find all IN_PROGRESS participants on email steps with wait_for_landing
            participants = _db().query(Participant).filter_by(
                status=ParticipantStatus.IN_PROGRESS
            ).all()

            checked = 0
            acted = 0

            for p in participants:
                step = p.current_step
                if not step or step.type.value != 'email':
                    continue
                config = step.skip_conditions or {}
                if not config.get('wait_for_landing'):
                    continue

                # Skip if participant has scheduled executions (still progressing)
                has_scheduled = _db().query(Execution).filter_by(
                    participant_id=p.id,
                    status=ExecutionStatus.SCHEDULED
                ).first()
                if has_scheduled:
                    continue

                checked += 1
                timeout_days = config.get('landing_timeout_days', 7)
                if_filled = config.get('landing_if_filled', 'continue')
                if_filled_step = config.get('landing_if_filled_step', 0)
                if_timeout = config.get('landing_if_timeout', 'continue')
                if_timeout_step = config.get('landing_if_timeout_step', 0)

                # Find when email was sent
                last_exec = _db().query(Execution).filter_by(
                    participant_id=p.id, step_id=step.id
                ).filter(
                    Execution.status.in_([
                        ExecutionStatus.SENT, ExecutionStatus.DELIVERED,
                        ExecutionStatus.OPENED, ExecutionStatus.CLICKED
                    ])
                ).order_by(Execution.sent_at.desc()).first()

                if not last_exec or not last_exec.sent_at:
                    continue

                # Check if form was submitted
                if p.collected_data and len(p.collected_data) > 0:
                    logger.info(f"✓ Landing form submitted by participant {p.id}")
                    SchedulerService._handle_landing_branch(p, step, if_filled, if_filled_step)
                    acted += 1
                    continue

                # Check timeout
                timeout_at = last_exec.sent_at + timedelta(days=timeout_days)
                if datetime.utcnow() >= timeout_at:
                    logger.info(f"⏰ Landing wait timeout for participant {p.id}")
                    SchedulerService._handle_landing_branch(p, step, if_timeout, if_timeout_step)
                    acted += 1

            if checked > 0:
                logger.info(f"🔄 Landing wait check: {checked} checked, {acted} acted on")

        except Exception as e:
            logger.error(f"✗ Landing wait cron error: {str(e)}")
            _db().rollback()

    @staticmethod
    def _handle_landing_branch(participant, step, action, jump_target):
        """Handle branching after landing wait (filled or timeout)."""
        if action == 'jump' and jump_target == 'end':
            action = 'stop'
        elif action == 'jump' and jump_target:
            target_step = _db().query(WorkflowStep).filter_by(
                workflow_id=step.workflow_id,
                order=jump_target
            ).first()
            if target_step:
                SchedulerService.schedule_step(participant, target_step, delay_hours=target_step.delay_hours)
                participant.current_step_id = target_step.id
                participant.status = ParticipantStatus.IN_PROGRESS
                _db().commit()
                logger.info(f"→ Jumped to step {target_step.order}: {target_step.name}")
                return
            logger.warning(f"⚠ Jump target step {jump_target} not found, continuing")
        elif action == 'stop':
            SchedulerService.cancel_scheduled_executions(participant.id)
            participant.status = ParticipantStatus.COMPLETED
            _db().commit()
            logger.info(f"✓ Workflow stopped for participant {participant.id}")
            return

        # Default: continue
        SchedulerService._schedule_next_step(participant, step)
        _db().commit()

    @staticmethod
    def _execute_email_step(participant, step, execution):
        """
        Execute email step
        
        Returns:
            bool: success
        """
        try:
            # Generate landing URL if step has landing config or template references it
            landing_url = None
            body = step.body_template or ''
            skip_cond = step.skip_conditions or {}
            has_landing = step.landing_page_config or skip_cond.get('has_landing') or '{{ landing_url }}' in body or '{{landing_url}}' in body
            if has_landing:
                landing_url = TokenService.generate_landing_url(participant)
            
            # Load attachments if configured
            attachments = []
            attachment_ids = skip_cond.get('attachment_ids', [])
            if attachment_ids:
                from app.models import Attachment
                attachments = _db().query(Attachment).filter(Attachment.id.in_(attachment_ids)).all()
                logger.info(f"Loaded {len(attachments)} attachments for email step")

            # Determine recipient
            to_override = None
            if skip_cond.get('recipient') == 'custom' and skip_cond.get('custom_to'):
                to_override = skip_cond['custom_to']

            # Send email
            success = EmailService.send_workflow_email(participant, step, landing_url, attachments=attachments, to_override=to_override)

            return success

        except Exception as e:
            logger.error(f"✗ Email error: {str(e)}")
            return False
    
    @staticmethod
    def _execute_human_approval_step(participant, step, execution):
        """Execute human approval step — sends email to approver with approve/reject buttons"""
        try:
            from flask import current_app
            from urllib.parse import quote
            config = step.skip_conditions or {}
            approver_email_raw = config.get('approver_email', '')
            approval_message = config.get('approval_message', '')

            # Parse multiple approver emails (comma or newline separated)
            approver_emails = [e.strip() for e in approver_email_raw.replace('\n', ',').split(',') if e.strip()]

            if not approver_emails:
                logger.error("✗ Human approval step missing approver_email")
                return False

            # Generate approval URL using participant token
            base_url = current_app.config.get('LANDING_BASE_URL', 'http://localhost:5001/landing')
            approval_base = base_url.rsplit('/landing', 1)[0] + '/approval'
            token = TokenService.generate_landing_url(participant)
            token_value = token.rsplit('/', 1)[-1]
            approve_url = f"{approval_base}/{token_value}?action=approve"
            reject_url = f"{approval_base}/{token_value}?action=reject"

            # Render approval message with context
            context = {
                'participant': {
                    'name': participant.full_name,
                    'first_name': participant.first_name,
                    'last_name': participant.last_name,
                    'email': participant.email,
                },
                'workflow_name': participant.workflow.name,
                'step_name': step.name,
            }
            if participant.workflow.config:
                context.update(participant.workflow.config)

            rendered_message = EmailService.render_template(approval_message, context) if approval_message else ''

            # Build participant data summary
            data_rows = ''
            # Base info
            base_fields = [
                ('Name', participant.full_name),
                ('Email', participant.email),
                ('Phone', participant.phone),
            ]
            for label, val in base_fields:
                if val:
                    data_rows += f'<tr><td style="padding:6px 12px;color:#888;font-size:13px;white-space:nowrap">{label}</td><td style="padding:6px 12px;font-size:13px">{val}</td></tr>'

            # Saba Form data
            sf = participant.sabaform_data
            if sf and isinstance(sf, dict):
                sf_labels = {'company': 'Company', 'birth_date': 'Birth Date', 'gender': 'Gender',
                             'nucleo': 'Family Unit', 'doc_type': 'Doc Type', 'doc_number': 'Doc Number',
                             'doc_expiry': 'Doc Expiry', 'volo_arrivo': 'Arrival Flight',
                             'volo_partenza': 'Departure Flight', 'notes': 'Notes'}
                for k, v in sf.items():
                    if v and k not in ('id', 'first_name', 'last_name', 'email', 'phone', 'confirmed', 'status'):
                        label = sf_labels.get(k, k)
                        data_rows += f'<tr><td style="padding:6px 12px;color:#888;font-size:13px;white-space:nowrap">{label}</td><td style="padding:6px 12px;font-size:13px">{v}</td></tr>'

            # Collected data (from landing page)
            cd = participant.collected_data
            if cd and isinstance(cd, dict):
                for k, v in cd.items():
                    if v and not isinstance(v, dict):
                        data_rows += f'<tr><td style="padding:6px 12px;color:#888;font-size:13px;white-space:nowrap">{k}</td><td style="padding:6px 12px;font-size:13px">{v}</td></tr>'

            data_table = ''
            if data_rows:
                data_table = f'<table style="width:100%;border-collapse:collapse;background:#fafafa;border-radius:8px;margin:16px 0">{data_rows}</table>'

            # Build email body
            body_html = f'''
            <div style="font-family: -apple-system, sans-serif; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #333;">Approval Required</h2>
                <p style="color: #666; font-size: 15px;">
                    Workflow: <strong>{participant.workflow.name}</strong>
                </p>
                {f'<div style="background: #fff3cd; padding: 16px; border-radius: 8px; margin: 16px 0; color: #333; border-left: 4px solid #ffc107;">{rendered_message}</div>' if rendered_message else ''}
                <h3 style="color:#555; font-size:16px; margin-top:24px;">Participant Data</h3>
                {data_table}
                <div style="text-align: center; margin-top: 30px; padding: 20px; background: #f5f5f5; border-radius: 12px;">
                    <a href="{approve_url}" style="display: inline-block; margin: 8px; padding: 14px 40px; background-color: #198754; color: #fff; text-decoration: none; border-radius: 8px; font-size: 16px; font-weight: 600;">Approve</a>
                    <a href="{reject_url}" style="display: inline-block; margin: 8px; padding: 14px 40px; background-color: #dc3545; color: #fff; text-decoration: none; border-radius: 8px; font-size: 16px; font-weight: 600;">Reject</a>
                </div>
                <p style="color: #999; font-size: 12px; margin-top: 20px; text-align: center;">
                    Timeout: {config.get('timeout_hours', 48)} hours
                </p>
            </div>
            '''

            subject = f"Approval Required: {participant.full_name or participant.email} — {participant.workflow.name}"

            # Send to all approvers
            sent_count = 0
            wf = participant.workflow
            for email in approver_emails:
                ok = EmailService.send_email(
                    to_email=email,
                    subject=subject,
                    body_html=body_html,
                    from_email=wf.mail_from_email or None,
                    from_name=wf.mail_from_name or None
                )
                if ok:
                    sent_count += 1

            logger.info(f"Approval email sent to {sent_count}/{len(approver_emails)} approvers")
            return sent_count > 0

        except Exception as e:
            logger.error(f"✗ Human approval error: {str(e)}")
            return False

    @staticmethod
    def _execute_survey_step(participant, step, execution):
        """Execute survey step — sends email with survey buttons"""
        try:
            from flask import current_app
            from urllib.parse import quote
            config = step.skip_conditions or {}
            question = config.get('question', '')
            response_type = config.get('response_type', 'choices')
            choices = config.get('choices', [])
            scale_max = config.get('scale_max', 5)

            # Generate survey URL (reuse token system)
            base_url = current_app.config.get('LANDING_BASE_URL', 'http://localhost:5001/landing')
            survey_base = base_url.rsplit('/landing', 1)[0] + '/survey'
            token = TokenService.generate_landing_url(participant)
            token_value = token.rsplit('/', 1)[-1]  # Extract just the token
            survey_url = f"{survey_base}/{token_value}"

            # Build choices list
            if response_type == 'scale':
                choices = [str(i) for i in range(1, scale_max + 1)]

            # Build survey buttons HTML
            buttons_html = '<div style="text-align:center; margin-top:30px; padding:20px; background:#f5f5f5; border-radius:12px;">'
            if question:
                buttons_html += f'<p style="font-size:18px; font-weight:600; margin-bottom:20px; color:#333;">{question}</p>'

            for choice in choices:
                encoded = quote(choice)
                btn_url = f"{survey_url}?choice={encoded}"
                buttons_html += (
                    f'<a href="{btn_url}" style="display:inline-block; margin:6px; '
                    f'padding:12px 28px; background-color:#795548; color:#fff; '
                    f'text-decoration:none; border-radius:8px; font-size:15px; '
                    f'font-weight:500;">{choice}</a> '
                )

            buttons_html += '</div>'

            # Render email body with context
            context = {
                'participant': {
                    'name': participant.full_name,
                    'first_name': participant.first_name,
                    'last_name': participant.last_name,
                    'email': participant.email,
                },
                'landing_url': '',
                'workflow_name': participant.workflow.name,
                'evento': participant.workflow.name,
                'step_name': step.name,
            }
            if participant.workflow.config:
                context.update(participant.workflow.config)

            body_html = EmailService.render_template(step.body_template or '', context)
            body_html += buttons_html

            subject = EmailService.render_template(step.subject or '', context)

            wf = participant.workflow
            success = EmailService.send_email(
                to_email=participant.email,
                subject=subject,
                body_html=body_html,
                from_email=wf.mail_from_email or None,
                from_name=wf.mail_from_name or None
            )

            return success

        except Exception as e:
            logger.error(f"✗ Survey error: {str(e)}")
            return False

    @staticmethod
    def _execute_goal_check_step(participant, step, execution):
        """
        Execute goal check step with jump/continue/complete/skip support.
        """
        try:
            config = step.skip_conditions or {}
            goal_type = config.get('goal', 'form_submitted')

            goal_met = SchedulerService._check_goal(participant, goal_type, config)

            action = config.get('if_met', 'complete') if goal_met else config.get('if_not_met', 'continue')
            jump_target = config.get('if_met_step', 0) if goal_met else config.get('if_not_met_step', 0)

            execution.result_data = {
                'goal_type': goal_type,
                'goal_met': goal_met,
                'action': action
            }

            logger.info(f"{'✓' if goal_met else '⊘'} Goal '{goal_type}' {'MET' if goal_met else 'NOT MET'} for participant {participant.id} → {action}")

            if action == 'complete':
                SchedulerService.cancel_scheduled_executions(participant.id)
                participant.status = ParticipantStatus.COMPLETED
                _db().commit()
                logger.info(f"✓ Workflow COMPLETED for participant {participant.id}")
                # Mark execution as sent but don't schedule next
                execution.status = ExecutionStatus.SENT
                execution.sent_at = datetime.utcnow()
                _db().commit()
                return True
            elif action == 'jump' and jump_target:
                target_step = _db().query(WorkflowStep).filter_by(
                    workflow_id=step.workflow_id,
                    order=jump_target
                ).first()
                if target_step:
                    SchedulerService.schedule_step(participant, target_step, delay_hours=target_step.delay_hours)
                    participant.current_step_id = target_step.id
                    participant.status = ParticipantStatus.IN_PROGRESS
                    execution.status = ExecutionStatus.SENT
                    execution.sent_at = datetime.utcnow()
                    _db().commit()
                    logger.info(f"→ Jumped to step {target_step.order}: {target_step.name}")
                    return True
                else:
                    logger.warning(f"⚠ Jump target step {jump_target} not found, continuing")
            elif action == 'skip':
                # Skip next step — schedule the one after
                next_order = step.order + 2
                skip_target = _db().query(WorkflowStep).filter_by(
                    workflow_id=step.workflow_id,
                    order=next_order
                ).first()
                if skip_target:
                    SchedulerService.schedule_step(participant, skip_target, delay_hours=skip_target.delay_hours)
                    participant.current_step_id = skip_target.id
                    execution.status = ExecutionStatus.SENT
                    execution.sent_at = datetime.utcnow()
                    _db().commit()
                    logger.info(f"→ Skipped next, jumping to step {next_order}: {skip_target.name}")
                    return True

            # Default: continue to next step
            return True

        except Exception as e:
            logger.error(f"✗ Goal check error: {str(e)}")
            return False
    
    @staticmethod
    def _check_goal(participant, goal_type, config):
        """
        Check if participant reached specific goal
        
        Args:
            participant: Participant instance
            goal_type: Type of goal to check
            config: Goal configuration
            
        Returns:
            bool: True if goal is met
        """
        try:
            if goal_type == 'form_submitted':
                # Check if participant submitted landing page form
                return participant.collected_data is not None and len(participant.collected_data) > 0
                
            elif goal_type == 'field_filled':
                # Check if specific field is filled
                field_name = config.get('field_name')
                if not field_name or not participant.collected_data:
                    return False
                return field_name in participant.collected_data and participant.collected_data[field_name]
                
            elif goal_type == 'field_equals':
                # Check if field equals specific value
                field_name = config.get('field_name')
                expected_value = config.get('field_value')
                if not field_name or not participant.collected_data:
                    return False
                return participant.collected_data.get(field_name) == expected_value
                
            elif goal_type == 'email_opened':
                # Check if participant opened any email (has interaction)
                return participant.last_interaction is not None
                
            elif goal_type == 'status_equals':
                # Check participant status
                expected_status = config.get('status_value', 'completed')
                return participant.status.value == expected_status
                
            else:
                logger.warning(f"Unknown goal type: {goal_type}")
                return False
                
        except Exception as e:
            logger.error(f"Error checking goal: {str(e)}")
            return False
    
    @staticmethod
    def _execute_export_data_step(participant, step, execution):
        """
        Execute export data step
        
        Exports all workflow participants to CSV and sends via email
        
        Returns:
            bool: success
        """
        try:
            from app.services.export_service import ExportService
            
            # Get export configuration from skip_conditions
            config = step.skip_conditions or {}
            export_format = config.get('format', 'csv')
            send_to = config.get('send_to', '')
            save_local = config.get('save_local', False)
            
            workflow_id = step.workflow_id
            
            logger.info(f"▶ Exporting workflow {workflow_id} data (format: {export_format})")
            
            if export_format == 'csv':
                success, csv_content, filename = ExportService.export_workflow_csv(
                    workflow_id=workflow_id,
                    send_to_email=send_to if send_to else None
                )
                
                if success:
                    execution.result_data = {
                        'format': export_format,
                        'filename': filename,
                        'sent_to': send_to,
                        'rows_exported': csv_content.count('\n') - 1 if csv_content else 0
                    }
                    
                    # Save locally if requested
                    if save_local and csv_content:
                        ExportService.save_csv_file(csv_content, filename)
                    
                    logger.info(f"✓ Export completed: {filename}")
                    return True
                else:
                    logger.error(f"✗ Export failed for workflow {workflow_id}")
                    return False
            else:
                logger.warning(f"Unsupported export format: {export_format}")
                return False
                
        except Exception as e:
            logger.error(f"✗ Export data error: {str(e)}")
            return False
    
    @staticmethod
    def _execute_condition_step(participant, step, execution):
        """
        Evaluate condition and decide branching.

        Reads field from sabaform_data, collected_data, or participant fields.
        Based on result, either continues, skips next step, or stops workflow.

        Returns:
            bool: True if evaluation succeeded (regardless of condition result)
        """
        try:
            from app.services.activity_service import log_activity

            config = step.skip_conditions or {}
            field_source = config.get('field_source', 'sabaform_data')
            field = config.get('field', '')
            operator = config.get('operator', 'equals')
            expected = config.get('value', '')
            if_true = config.get('if_true', 'continue')
            if_true_step = config.get('if_true_step', 0)
            if_false = config.get('if_false', 'continue')
            if_false_step = config.get('if_false_step', 0)

            # Get actual value from the right source
            actual = None
            if field_source == 'sabaform_data':
                actual = (participant.sabaform_data or {}).get(field)
            elif field_source == 'collected_data':
                actual = (participant.collected_data or {}).get(field)
            elif field_source == 'participant':
                actual = getattr(participant, field, None)
                if hasattr(actual, 'value'):  # Enum (e.g. status)
                    actual = actual.value

            actual_str = str(actual or '').strip()
            expected_str = str(expected or '').strip()

            # Evaluate condition
            if operator == 'equals':
                result = actual_str.lower() == expected_str.lower()
            elif operator == 'not_equals':
                result = actual_str.lower() != expected_str.lower()
            elif operator == 'contains':
                result = expected_str.lower() in actual_str.lower()
            elif operator == 'not_empty':
                result = bool(actual_str)
            elif operator == 'empty':
                result = not bool(actual_str)
            elif operator == 'greater_than':
                try:
                    result = float(actual_str) > float(expected_str)
                except (ValueError, TypeError):
                    result = actual_str > expected_str
            elif operator == 'less_than':
                try:
                    result = float(actual_str) < float(expected_str)
                except (ValueError, TypeError):
                    result = actual_str < expected_str
            else:
                logger.warning(f"Unknown operator: {operator}")
                result = False

            action = if_true if result else if_false

            execution.result_data = {
                'field': field,
                'field_source': field_source,
                'operator': operator,
                'expected': expected,
                'actual': actual_str,
                'result': result,
                'action': action,
            }

            logger.info(f"✓ Condition: {field} ({actual_str}) {operator} {expected_str} → {result} → {action}")

            log_activity(
                workflow_id=step.workflow_id,
                event_type='condition_evaluated',
                description=f'Condizione "{field} {operator} {expected}": {"VERO" if result else "FALSO"} → {action}',
                participant_id=participant.id,
                step_id=step.id,
                details=execution.result_data
            )

            # Determine target step order for jump
            jump_target = if_true_step if result else if_false_step

            # Execute the action
            if action == 'stop':
                SchedulerService.cancel_scheduled_executions(participant.id)
                participant.status = ParticipantStatus.COMPLETED
                _db().commit()
                logger.info(f"✓ Workflow STOPPED for participant {participant.id} by condition")
            elif action == 'jump' and jump_target:
                # Jump to specific step
                target_step = _db().query(WorkflowStep).filter_by(
                    workflow_id=step.workflow_id,
                    order=jump_target
                ).first()
                if target_step:
                    SchedulerService.schedule_step(participant, target_step, delay_hours=target_step.delay_hours)
                    participant.current_step_id = target_step.id
                    participant.status = ParticipantStatus.IN_PROGRESS
                    _db().commit()
                    logger.info(f"→ Jumped to step {target_step.order}: {target_step.name}")
                else:
                    logger.warning(f"⚠ Jump target step order {jump_target} not found, continuing normally")
                    SchedulerService._schedule_next_step(participant, step)
            else:
                # continue — schedule next step normally
                SchedulerService._schedule_next_step(participant, step)

            return True

        except Exception as e:
            logger.error(f"✗ Condition error: {str(e)}")
            return False

    @staticmethod
    def _execute_whatsapp_step(participant, step, execution):
        """Send a WhatsApp message via Meta Business API."""
        try:
            from flask import current_app
            import requests as http_requests
            from jinja2 import Template

            config = step.skip_conditions or {}
            phone = participant.phone or ''
            if not phone:
                logger.warning(f"⚠ WhatsApp step skipped: no phone for participant {participant.id}")
                execution.result_data = {'error': 'no_phone'}
                return False

            # Normalize phone: ensure starts with +, remove spaces/dashes
            phone = phone.strip().replace(' ', '').replace('-', '')
            if not phone.startswith('+'):
                phone = '+' + phone
            # WhatsApp API wants digits only, no +
            phone_digits = phone.lstrip('+')

            phone_id = current_app.config.get('WHATSAPP_PHONE_ID', '')
            token = current_app.config.get('WHATSAPP_TOKEN', '')
            if not phone_id or not token:
                logger.error("WhatsApp step: WHATSAPP_PHONE_ID or WHATSAPP_TOKEN not configured")
                return False

            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            url = f"https://graph.facebook.com/v21.0/{phone_id}/messages"

            message_type = config.get('message_type', 'template')

            if message_type == 'text':
                # Free text message (only within 24h conversation window)
                body_text = config.get('body_text', '')
                if body_text:
                    # Render template variables
                    try:
                        tpl = Template(body_text)
                        body_text = tpl.render(
                            participant=participant,
                            workflow_name=step.workflow.name if step.workflow else '',
                            step_name=step.name
                        )
                    except Exception:
                        pass

                payload = {
                    "messaging_product": "whatsapp",
                    "to": phone_digits,
                    "type": "text",
                    "text": {"body": body_text}
                }
            else:
                # Template message
                template_name = config.get('template_name', 'hello_world')
                template_lang = config.get('template_language', 'en_US')
                payload = {
                    "messaging_product": "whatsapp",
                    "to": phone_digits,
                    "type": "template",
                    "template": {
                        "name": template_name,
                        "language": {"code": template_lang}
                    }
                }

            resp = http_requests.post(url, headers=headers, json=payload, timeout=15)

            if resp.status_code in (200, 201):
                resp_data = resp.json()
                msg_id = resp_data.get('messages', [{}])[0].get('id', '')
                logger.info(f"✓ WhatsApp sent to {phone_digits} (msg: {msg_id})")
                execution.result_data = {'phone': phone_digits, 'message_id': msg_id, 'type': message_type}
                return True

            logger.error(f"WhatsApp send failed ({resp.status_code}): {resp.text}")
            execution.result_data = {'error': resp.text}
            return False

        except Exception as e:
            logger.error(f"✗ WhatsApp error: {str(e)}")
            return False

    @staticmethod
    def _execute_excel_write_step(participant, step, execution):
        """Write a row to an Excel file (OneDrive, SharePoint, or local)."""
        try:
            from flask import current_app

            config = step.skip_conditions or {}
            file_path = config.get('file_path', '')
            sheet_name = config.get('sheet_name', 'Sheet1')
            columns = config.get('columns', [])
            storage = config.get('storage', 'onedrive')

            if not file_path:
                execution.error_message = "file_path not configured"
                logger.error(f"Excel write step: {execution.error_message}")
                return False

            # Auto-generate columns from collected_data if none configured
            if not columns:
                cd = participant.collected_data or {}
                columns = [{'header': k, 'source': 'collected_data', 'field': k} for k in cd.keys()]
                if not columns:
                    execution.error_message = "no columns configured and no collected_data"
                    logger.error(f"Excel write step: {execution.error_message}")
                    return False
                logger.info(f"Excel write: auto-generated {len(columns)} columns from collected_data")

            # Build row values from participant data
            collected = participant.collected_data or {}
            sabaform = participant.sabaform_data or {}

            # Helper: find value in a dict with case-insensitive/fuzzy fallback
            def _lookup(data, field):
                # Exact match
                if field in data:
                    return data[field]
                # Case-insensitive match
                field_lower = field.lower().replace(' ', '_').replace("'", '')
                for k, v in data.items():
                    if k.lower().replace(' ', '_').replace("'", '') == field_lower:
                        return v
                return None

            # If ALL columns are collected_data and none match, use positional mapping
            all_collected = all(c.get('source', 'participant') == 'collected_data' for c in columns)
            collected_values = list(collected.values()) if collected else []
            # Check if any configured field actually matches a collected_data key
            any_match = any(_lookup(collected, c.get('field', '')) is not None for c in columns if c.get('source') == 'collected_data')

            row_values = []
            for i, col in enumerate(columns):
                field = col.get('field', '')
                source = col.get('source', 'participant')
                val = ''
                if source == 'participant':
                    if field in ('created_at', 'last_interaction'):
                        dt = getattr(participant, field, None)
                        val = dt.strftime('%Y-%m-%d %H:%M') if dt else ''
                    else:
                        val = getattr(participant, field, '')
                        if hasattr(val, 'value'):
                            val = val.value
                elif source == 'collected_data':
                    found = _lookup(collected, field)
                    if found is not None:
                        val = found
                    elif all_collected and not any_match and i < len(collected_values):
                        # Positional fallback: map columns to collected_data values in order
                        val = collected_values[i]
                elif source == 'sabaform_data':
                    found = _lookup(sabaform, field)
                    val = found if found is not None else ''
                row_values.append(str(val or ''))

            # === LOCAL FILE ===
            if storage == 'local':
                return SchedulerService._excel_write_local(file_path, sheet_name, row_values, columns, execution)

            # === GRAPH API (OneDrive / SharePoint) ===
            from app.services.email_service import EmailService
            import requests as http_requests

            token = EmailService._get_access_token()
            from_email = current_app.config.get('MAIL_FROM_EMAIL', '')
            headers = {
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json"
            }
            clean_path = file_path.strip('/')

            if storage == 'sharepoint':
                # SharePoint: resolve site then use its drive
                sp_site = config.get('sharepoint_site', '').strip().strip('/')
                if not sp_site:
                    execution.error_message = "SharePoint site not configured"
                    logger.error(f"Excel write: {execution.error_message}")
                    return False
                site_url = f"https://graph.microsoft.com/v1.0/sites/{sp_site}"
                site_resp = http_requests.get(site_url, headers=headers, timeout=15)
                if site_resp.status_code != 200:
                    execution.error_message = f"SharePoint site not found: {site_resp.text[:200]}"
                    logger.error(f"Excel write: {execution.error_message}")
                    return False
                site_id = site_resp.json()['id']
                drive_base = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive"
            else:
                # OneDrive personal
                drive_base = f"https://graph.microsoft.com/v1.0/users/{from_email}/drive"

            # Get file item by path
            file_url = f"{drive_base}/root:/{clean_path}"
            file_resp = http_requests.get(file_url, headers=headers, timeout=15)
            if file_resp.status_code != 200:
                execution.error_message = f"File not found at {file_path}: {file_resp.text[:200]}"
                logger.error(f"Excel write: {execution.error_message}")
                return False

            file_id = file_resp.json()['id']
            wb_base = f"{drive_base}/items/{file_id}/workbook"

            # Create workbook session to get edit access
            session_url = f"{wb_base}/createSession"
            session_resp = http_requests.post(session_url, headers=headers,
                                              json={"persistChanges": True}, timeout=15)
            if session_resp.status_code in (200, 201):
                session_id = session_resp.json().get('id')
                headers["workbook-session-id"] = session_id
                logger.info(f"Excel write: workbook session created")
            else:
                logger.warning(f"Excel write: session creation failed ({session_resp.status_code}), proceeding without session")

            # Resolve actual sheet name (configured name may differ from real name, e.g. Sheet1 vs Foglio1)
            sheets_resp = http_requests.get(f"{wb_base}/worksheets", headers=headers, timeout=15)
            if sheets_resp.status_code == 200:
                sheets = sheets_resp.json().get('value', [])
                # Try to match configured name first, otherwise use first sheet
                real_sheet = None
                for sh in sheets:
                    if sh['name'].lower() == sheet_name.lower():
                        real_sheet = sh['name']
                        break
                if not real_sheet and sheets:
                    real_sheet = sheets[0]['name']
                    logger.info(f"Excel write: sheet '{sheet_name}' not found, using '{real_sheet}'")
                if real_sheet:
                    sheet_name = real_sheet

            workbook_base = f"{wb_base}/worksheets('{sheet_name}')"

            # Try table first
            add_url = f"{workbook_base}/tables/@/rows/add"
            payload = {"values": [row_values]}
            resp = http_requests.post(add_url, headers=headers, json=payload, timeout=15)

            if resp.status_code in (200, 201):
                logger.info(f"✓ Excel row added to {file_path} ({len(columns)} columns)")
                execution.result_data = {'file': file_path, 'storage': storage, 'columns': len(columns)}
                return True

            # Fallback: append to used range
            range_url = f"{workbook_base}/usedRange"
            range_resp = http_requests.get(range_url, headers=headers, timeout=15)
            if range_resp.status_code != 200:
                execution.error_message = f"Cannot read used range: {range_resp.text[:200]}"
                logger.error(f"Excel write: {execution.error_message}")
                return False

            next_row = range_resp.json().get('rowCount', 1) + 1
            last_col_letter = chr(ord('A') + len(row_values) - 1) if len(row_values) <= 26 else 'Z'
            cell_range = f"A{next_row}:{last_col_letter}{next_row}"

            patch_url = f"{workbook_base}/range(address='{cell_range}')"
            patch_resp = http_requests.patch(patch_url, headers=headers, json={"values": [row_values]}, timeout=15)

            if patch_resp.status_code == 200:
                logger.info(f"✓ Excel row written at {cell_range} in {file_path}")
                execution.result_data = {'file': file_path, 'storage': storage, 'range': cell_range}
                return True

            execution.error_message = f"Graph API error {patch_resp.status_code}: {patch_resp.text[:200]}"
            logger.error(f"Excel write failed: {execution.error_message}")
            return False

        except Exception as e:
            execution.error_message = f"Excel write error: {str(e)}"
            logger.error(f"✗ {execution.error_message}")
            return False
        finally:
            # Close workbook session
            if headers.get('workbook-session-id'):
                try:
                    http_requests.post(f"{wb_base}/closeSession", headers=headers, timeout=10)
                except Exception:
                    pass

    @staticmethod
    def _excel_write_local(file_path, sheet_name, row_values, columns, execution):
        """Write a row to a local Excel file using openpyxl."""
        try:
            import openpyxl
            import os

            if os.path.exists(file_path):
                wb = openpyxl.load_workbook(file_path)
            else:
                # Create new file with headers
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = sheet_name
                headers = [col.get('header', f'Col{i+1}') for i, col in enumerate(columns)]
                ws.append(headers)

            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                headers = [col.get('header', f'Col{i+1}') for i, col in enumerate(columns)]
                ws.append(headers)

            ws.append(row_values)
            wb.save(file_path)
            wb.close()

            logger.info(f"✓ Excel row written locally to {file_path} ({len(row_values)} columns)")
            execution.result_data = {'file': file_path, 'storage': 'local', 'row': ws.max_row}
            return True

        except Exception as e:
            logger.error(f"✗ Local Excel write error: {str(e)}")
            return False

    @staticmethod
    def _should_skip(participant, step):
        """
        Check if step should be skipped based on conditions
        
        Args:
            participant: Participant instance
            step: WorkflowStep instance
            
        Returns:
            bool: True if should skip
        """
        # TODO: Implement JSONLogic for skip_conditions
        # For now, always returns False
        return False
    
    @staticmethod
    def _schedule_next_step(participant, current_step):
        """
        Schedule next step — uses explicit next_step config if set,
        otherwise falls back to sequential order.
        """
        try:
            config = current_step.skip_conditions or {}
            explicit_next = config.get('next_step')

            next_step = None

            if explicit_next == 'end':
                # Explicit END
                logger.info(f"✓ Workflow completed for participant {participant.id} (explicit END)")
                participant.status = ParticipantStatus.COMPLETED
                _db().commit()
                return
            elif explicit_next and explicit_next != 'auto':
                # Jump to specific step by order
                try:
                    target_order = int(explicit_next)
                    next_step = _db().query(WorkflowStep).filter_by(
                        workflow_id=current_step.workflow_id,
                        order=target_order
                    ).first()
                    if next_step:
                        logger.info(f"→ Explicit next_step: jumping to step {target_order}")
                except (ValueError, TypeError):
                    pass

            if not next_step:
                # Fallback: sequential order
                next_step = _db().query(WorkflowStep).filter_by(
                    workflow_id=current_step.workflow_id,
                    order=current_step.order + 1
                ).first()

            if next_step:
                SchedulerService.schedule_step(
                    participant,
                    next_step,
                    delay_hours=next_step.delay_hours
                )
                participant.current_step_id = next_step.id
                participant.status = ParticipantStatus.IN_PROGRESS
                _db().commit()
                logger.info(f"→ Scheduled next step {next_step.id} for participant {participant.id}")
            else:
                logger.info(f"✓ Workflow completed for participant {participant.id} (no more steps)")
                participant.status = ParticipantStatus.COMPLETED
                _db().commit()

        except Exception as e:
            logger.error(f"✗ Error scheduling next step: {str(e)}")

    @staticmethod
    def reconcile_participant(participant_id):
        """
        Re-evaluate a stuck participant and resume their workflow.
        Re-reads current config from DB (not cached thread state).
        Returns dict with action taken.
        """
        from app.models import Participant
        p = _db().get(Participant, participant_id)
        if not p or p.status != ParticipantStatus.IN_PROGRESS:
            return {'action': 'skipped', 'reason': 'not_in_progress'}

        # Check if already has scheduled executions (not stuck)
        has_scheduled = _db().query(Execution).filter_by(
            participant_id=p.id,
            status=ExecutionStatus.SCHEDULED
        ).first()
        if has_scheduled:
            return {'action': 'skipped', 'reason': 'already_scheduled'}

        step = p.current_step
        if not step:
            return {'action': 'skipped', 'reason': 'no_current_step'}

        config = step.skip_conditions or {}

        # Email step with wait_for_landing
        if step.type.value == 'email' and config.get('wait_for_landing'):
            timeout_days = config.get('landing_timeout_days', 7)
            if_filled = config.get('landing_if_filled', 'continue')
            if_filled_step = config.get('landing_if_filled_step', 0)
            if_timeout = config.get('landing_if_timeout', 'continue')
            if_timeout_step = config.get('landing_if_timeout_step', 0)

            # Find when email was sent
            last_exec = _db().query(Execution).filter_by(
                participant_id=p.id,
                step_id=step.id,
            ).filter(
                Execution.status.in_([
                    ExecutionStatus.SENT, ExecutionStatus.DELIVERED,
                    ExecutionStatus.OPENED, ExecutionStatus.CLICKED
                ])
            ).order_by(Execution.sent_at.desc()).first()

            # Form already submitted?
            if p.collected_data and len(p.collected_data) > 0:
                SchedulerService._handle_landing_branch(p, step, if_filled, if_filled_step)
                return {'action': 'form_filled_branch', 'branch': if_filled,
                        'target': if_filled_step if if_filled == 'jump' else if_filled}

            # Timeout passed?
            if last_exec and last_exec.sent_at:
                timeout_at = last_exec.sent_at + timedelta(days=timeout_days)
                if datetime.utcnow() >= timeout_at:
                    SchedulerService._handle_landing_branch(p, step, if_timeout, if_timeout_step)
                    return {'action': 'timeout_branch', 'branch': if_timeout,
                            'target': if_timeout_step if if_timeout == 'jump' else if_timeout}

            # Still within timeout — cron job will pick it up automatically
            return {'action': 'waiting', 'reason': 'within_timeout'}

        # Non-landing step: just schedule next step
        SchedulerService._schedule_next_step(p, step)
        return {'action': 'next_step_scheduled'}

    @staticmethod
    def get_reconcile_status(workflow_id):
        """
        Get breakdown of all participants by step and status for reconciliation.
        """
        from app.models import Participant, Workflow
        workflow = _db().get(Workflow, workflow_id)
        if not workflow:
            return None

        participants = _db().query(Participant).filter_by(workflow_id=workflow_id).all()

        # Find participants with scheduled executions
        scheduled_pids = set(
            r[0] for r in _db().query(Execution.participant_id).filter(
                Execution.participant_id.in_([p.id for p in participants]),
                Execution.status == ExecutionStatus.SCHEDULED
            ).all()
        ) if participants else set()

        summary = {'total': len(participants), 'completed': 0, 'pending': 0,
                   'in_progress': 0, 'waiting_landing': 0, 'actionable': 0,
                   'bounced': 0, 'unsubscribed': 0}
        by_step = {}

        for p in participants:
            if p.status == ParticipantStatus.COMPLETED:
                summary['completed'] += 1
                continue
            elif p.status == ParticipantStatus.PENDING:
                summary['pending'] += 1
                continue
            elif p.status == ParticipantStatus.BOUNCED:
                summary['bounced'] += 1
                continue
            elif p.status == ParticipantStatus.UNSUBSCRIBED:
                summary['unsubscribed'] += 1
                continue

            # IN_PROGRESS — determine sub-state
            step = p.current_step
            if not step:
                summary['in_progress'] += 1
                continue

            config = step.skip_conditions or {}
            is_landing_wait = (step.type.value == 'email' and config.get('wait_for_landing'))

            if not is_landing_wait:
                if p.id in scheduled_pids:
                    summary['in_progress'] += 1
                else:
                    summary['actionable'] += 1
                    # Group stuck non-landing
                    step_key = step.id
                    if step_key not in by_step:
                        by_step[step_key] = {
                            'step_id': step.id, 'step_order': step.order,
                            'step_name': step.name, 'step_type': step.type.value,
                            'wait_for_landing': False, 'participants': []
                        }
                    by_step[step_key]['participants'].append({
                        'id': p.id,
                        'name': ((p.first_name or '') + ' ' + (p.last_name or '')).strip() or p.email,
                        'email': p.email, 'status_label': 'stuck',
                        'has_form_data': False, 'timeout_passed': False,
                        'email_sent': False, 'exec_status': None, 'sent_at': None
                    })
                continue

            # Landing wait — get execution context
            last_exec = _db().query(Execution).filter_by(
                participant_id=p.id, step_id=step.id
            ).filter(
                Execution.status.in_([
                    ExecutionStatus.SENT, ExecutionStatus.DELIVERED,
                    ExecutionStatus.OPENED, ExecutionStatus.CLICKED
                ])
            ).order_by(Execution.sent_at.desc()).first()

            has_form = bool(p.collected_data and len(p.collected_data) > 0)
            email_sent = bool(last_exec and last_exec.sent_at)
            timeout_days = config.get('landing_timeout_days', 7)
            timeout_passed = False
            if email_sent:
                timeout_passed = datetime.utcnow() >= (last_exec.sent_at + timedelta(days=timeout_days))

            # Determine sub-state
            if has_form or timeout_passed:
                summary['actionable'] += 1
                status_label = 'form_compilato' if has_form else 'timeout_scaduto'
            else:
                summary['waiting_landing'] += 1
                status_label = 'in_attesa' if email_sent else 'mail_non_inviata'

            # Group by step
            step_key = step.id
            if step_key not in by_step:
                by_step[step_key] = {
                    'step_id': step.id, 'step_order': step.order,
                    'step_name': step.name, 'step_type': step.type.value,
                    'wait_for_landing': True, 'participants': []
                }

            by_step[step_key]['participants'].append({
                'id': p.id,
                'name': ((p.first_name or '') + ' ' + (p.last_name or '')).strip() or p.email,
                'email': p.email,
                'status_label': status_label,
                'has_form_data': has_form,
                'timeout_passed': timeout_passed,
                'email_sent': email_sent,
                'exec_status': last_exec.status.value if last_exec else None,
                'sent_at': last_exec.sent_at.isoformat() if email_sent else None
            })

        return {
            'workflow_id': workflow_id,
            'workflow_name': workflow.name,
            'summary': summary,
            'by_step': sorted(by_step.values(), key=lambda x: x['step_order'])
        }

    @staticmethod
    def cancel_scheduled_executions(participant_id):
        """
        Cancel all scheduled executions for a participant
        
        Args:
            participant_id: Participant ID
        """
        try:
            executions = _db().query(Execution).filter_by(
                participant_id=participant_id,
                status=ExecutionStatus.SCHEDULED
            ).all()
            
            for execution in executions:
                if execution.job_id:
                    try:
                        _app.scheduler.remove_job(execution.job_id)
                        logger.info(f"✓ Cancelled job {execution.job_id}")
                    except Exception as e:
                        logger.warning(f"⚠ Job {execution.job_id} not found: {str(e)}")
                
                execution.status = ExecutionStatus.SKIPPED
                execution.result_data = {'reason': 'Cancelled by user action'}
            
            _db().commit()
            logger.info(f"✓ Cancelled {len(executions)} executions for participant {participant_id}")
            
        except Exception as e:
            logger.error(f"✗ Error cancelling executions: {str(e)}")
            _db().rollback()
