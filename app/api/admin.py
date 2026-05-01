from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, Response
from app import db_session as db
from app.models import Workflow, WorkflowStep, Participant, Execution, ActivityLog, WorkflowStatus, ParticipantStatus, ExecutionStatus, UploadedImage, Attachment, User, user_workflows
from app.api.auth import superuser_required
from sqlalchemy import func
from sqlalchemy.orm import joinedload, selectinload, subqueryload
from datetime import datetime
import logging
import uuid

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def get_visible_workflows(query=None):
    """Filter workflows based on current user permissions"""
    from flask import g
    if query is None:
        query = db.query(Workflow)
    user = getattr(g, 'user', None)
    if user and user.is_superuser:
        return query
    if user:
        wf_ids = [wf.id for wf in user.workflows]
        return query.filter(Workflow.id.in_(wf_ids)) if wf_ids else query.filter(Workflow.id == -1)
    return query.filter(Workflow.id == -1)


@admin_bp.route('/')
@admin_bp.route('/dashboard')
def dashboard():
    """Dashboard principale"""
    try:
        # Statistiche generali
        stats = {
            'total_workflows': db.query(Workflow).count(),
            'active_workflows': db.query(Workflow).filter_by(status=WorkflowStatus.ACTIVE).count(),
            'total_participants': db.query(Participant).count(),
            'emails_sent': db.query(Execution).filter_by(status=ExecutionStatus.SENT).count()
        }
        
        # Workflows per stato
        workflow_status_query = db.query(
            Workflow.status,
            func.count(Workflow.id)
        ).group_by(Workflow.status).all()
        
        workflow_status_labels = [status.value for status, _ in workflow_status_query]
        workflow_status_data = [count for _, count in workflow_status_query]
        
        # Partecipanti per stato
        participant_status_query = db.query(
            Participant.status,
            func.count(Participant.id)
        ).group_by(Participant.status).all()
        
        participant_status_labels = [status.value for status, _ in participant_status_query]
        participant_status_data = [count for _, count in participant_status_query]
        
        # Workflows recenti (filtered by user)
        recent_workflows = get_visible_workflows().options(
            selectinload(Workflow.steps),
            selectinload(Workflow.participants)
        ).order_by(
            Workflow.created_at.desc()
        ).limit(5).all()

        # Attività recenti (log unificato)
        recent_activities = db.query(ActivityLog).options(
            joinedload(ActivityLog.workflow),
            joinedload(ActivityLog.participant)
        ).order_by(
            ActivityLog.created_at.desc()
        ).limit(10).all()

        # Engagement — participants with last event (optimized: 3 queries instead of N*2)
        engagement_workflows = get_visible_workflows().options(
            selectinload(Workflow.participants)
        ).filter(Workflow.participants.any()).all()

        # Batch-fetch last activity and last execution per participant via subqueries
        participant_ids = [p.id for wf in engagement_workflows for p in wf.participants]

        last_activities = {}
        last_executions = {}
        if participant_ids:
            # Latest ActivityLog per participant
            latest_act_sub = db.query(
                ActivityLog.participant_id,
                func.max(ActivityLog.id).label('max_id')
            ).filter(
                ActivityLog.participant_id.in_(participant_ids)
            ).group_by(ActivityLog.participant_id).subquery()

            for a in db.query(ActivityLog).join(
                latest_act_sub, ActivityLog.id == latest_act_sub.c.max_id
            ).all():
                last_activities[a.participant_id] = a

            # Latest Execution per participant
            latest_exec_sub = db.query(
                Execution.participant_id,
                func.max(Execution.id).label('max_id')
            ).filter(
                Execution.participant_id.in_(participant_ids)
            ).group_by(Execution.participant_id).subquery()

            for ex in db.query(Execution).join(
                latest_exec_sub, Execution.id == latest_exec_sub.c.max_id
            ).all():
                last_executions[ex.participant_id] = ex

        engagement_participants = []
        for wf in engagement_workflows:
            for p in wf.participants:
                last_activity = last_activities.get(p.id)
                last_execution = last_executions.get(p.id)

                last_event = None
                last_event_time = None
                if last_activity and last_execution:
                    if (last_activity.created_at or datetime.min) > (last_execution.created_at or datetime.min):
                        last_event = last_activity.event_type
                        last_event_time = last_activity.created_at
                    else:
                        last_event = last_execution.status.value if last_execution.status else 'unknown'
                        last_event_time = last_execution.sent_at or last_execution.created_at
                elif last_activity:
                    last_event = last_activity.event_type
                    last_event_time = last_activity.created_at
                elif last_execution:
                    last_event = last_execution.status.value if last_execution.status else 'unknown'
                    last_event_time = last_execution.sent_at or last_execution.created_at

                engagement_participants.append({
                    'id': p.id,
                    'name': p.full_name or p.email,
                    'email': p.email,
                    'status': p.status.value,
                    'workflow_id': wf.id,
                    'workflow_name': wf.name,
                    'last_event': last_event,
                    'last_event_time': last_event_time,
                })

        return render_template('admin/dashboard.html',
                             stats=stats,
                             workflow_status_labels=workflow_status_labels,
                             workflow_status_data=workflow_status_data,
                             participant_status_labels=participant_status_labels,
                             participant_status_data=participant_status_data,
                             recent_workflows=recent_workflows,
                             recent_activities=recent_activities,
                             engagement_workflows=engagement_workflows,
                             engagement_participants=engagement_participants)

    except Exception as e:
        logger.error(f"Errore dashboard: {str(e)}")
        flash(f'Errore caricamento dashboard: {str(e)}', 'danger')
        return render_template('admin/dashboard.html',
                             stats={'total_workflows': 0, 'active_workflows': 0,
                                   'total_participants': 0, 'emails_sent': 0},
                             workflow_status_labels=[],
                             workflow_status_data=[],
                             participant_status_labels=[],
                             participant_status_data=[],
                             recent_workflows=[],
                             recent_activities=[],
                             engagement_workflows=[],
                             engagement_participants=[])


@admin_bp.route('/workflows')
def workflows_list():
    """Lista tutti i workflows"""
    try:
        # Filtri
        status_filter = request.args.get('status')
        
        query = get_visible_workflows()

        if status_filter:
            query = query.filter_by(status=WorkflowStatus(status_filter))

        workflows = query.order_by(Workflow.created_at.desc()).all()
        
        return render_template('admin/workflows_list.html',
                             workflows=workflows,
                             current_filter=status_filter)
    
    except Exception as e:
        logger.error(f"Errore lista workflows: {str(e)}")
        flash(f'Errore caricamento workflows: {str(e)}', 'danger')
        return render_template('admin/workflows_list.html',
                             workflows=[],
                             current_filter=None)


@admin_bp.route('/workflows/create')
def workflow_create():
    """Form creazione workflow"""
    return render_template('admin/workflow_form.html',
                         workflow=None,
                         mode='create')


@admin_bp.route('/workflows/<int:workflow_id>')
def workflow_detail(workflow_id):
    """Dettaglio workflow"""
    try:
        workflow = db.get(Workflow, workflow_id)
        
        if not workflow:
            flash('Workflow non trovato', 'danger')
            return redirect(url_for('admin.workflows_list'))
        
        return render_template('admin/workflow_detail.html',
                             workflow=workflow)
    
    except Exception as e:
        logger.error(f"Errore dettaglio workflow: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return redirect(url_for('admin.workflows_list'))


@admin_bp.route('/workflows/<int:workflow_id>/edit')
def workflow_edit(workflow_id):
    """Form modifica workflow"""
    try:
        workflow = db.get(Workflow, workflow_id)
        
        if not workflow:
            flash('Workflow non trovato', 'danger')
            return redirect(url_for('admin.workflows_list'))
        
        return render_template('admin/workflow_form.html',
                             workflow=workflow,
                             mode='edit')
    
    except Exception as e:
        logger.error(f"Errore modifica workflow: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return redirect(url_for('admin.workflows_list'))


@admin_bp.route('/participants')
def participants_list():
    """Lista partecipanti"""
    try:
        workflow_id = request.args.get('workflow_id', type=int)
        LIMIT = 500

        workflows = db.query(Workflow).all()

        # Default to most recently updated active workflow
        if not workflow_id and workflows:
            active = [w for w in workflows if w.status.value == 'active']
            if active:
                workflow_id = max(active, key=lambda w: w.updated_at or w.created_at).id
            else:
                workflow_id = max(workflows, key=lambda w: w.updated_at or w.created_at).id

        return render_template('admin/participants_list.html',
                             workflows=workflows,
                             current_workflow_id=workflow_id)

    except Exception as e:
        logger.error(f"Errore lista partecipanti: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return render_template('admin/participants_list.html',
                             workflows=[],
                             current_workflow_id=None)


@admin_bp.route('/api/participants-list')
def participants_list_api():
    """API endpoint for participants list (AJAX, lightweight)"""
    import pytz
    local_tz = pytz.timezone('Europe/Rome')

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        LIMIT = 500

        wf_names = {w.id: w.name for w in db.query(Workflow.id, Workflow.name).all()}

        query = db.query(
            Participant.id, Participant.first_name, Participant.last_name,
            Participant.email, Participant.phone, Participant.workflow_id,
            Participant.status, Participant.enrolled_at, Participant.last_interaction,
            Participant.completed_at, Participant.token,
            Participant.current_step_id,
            WorkflowStep.name.label('step_name'),
            WorkflowStep.order.label('step_order'),
        ).outerjoin(WorkflowStep, Participant.current_step_id == WorkflowStep.id)

        if workflow_id:
            query = query.filter(Participant.workflow_id == workflow_id)

        rows = query.order_by(Participant.enrolled_at.desc()).limit(LIMIT).all()

        # Collect unique keys from collected_data/sabaform_data (lightweight query)
        cd_keys = set()
        sf_keys = set()
        key_query = db.query(Participant.collected_data, Participant.sabaform_data)
        if workflow_id:
            key_query = key_query.filter(Participant.workflow_id == workflow_id)
        for cd, sf in key_query.filter(
            (Participant.collected_data.isnot(None)) | (Participant.sabaform_data.isnot(None))
        ).limit(LIMIT).all():
            if cd and isinstance(cd, dict):
                cd_keys.update(cd.keys())
            if sf and isinstance(sf, dict):
                sf_keys.update(sf.keys())

        def _fmt(dt):
            if dt is None:
                return None
            utc_dt = pytz.utc.localize(dt)
            return utc_dt.astimezone(local_tz).strftime('%d/%m/%Y %H:%M')

        participants = []
        for row in rows:
            participants.append({
                'id': row.id,
                'first_name': row.first_name or '',
                'last_name': row.last_name or '',
                'email': row.email or '',
                'phone': row.phone or '',
                'workflow': wf_names.get(row.workflow_id, ''),
                'workflow_id': row.workflow_id,
                'status': row.status.value,
                'current_step': row.step_name or '',
                'current_step_order': row.step_order,
                'enrolled_at': _fmt(row.enrolled_at),
                'last_interaction': _fmt(row.last_interaction),
                'completed_at': _fmt(row.completed_at),
                'token': row.token or '',
            })

        return jsonify({
            'participants': participants,
            'cd_keys': sorted(cd_keys),
            'sf_keys': sorted(sf_keys),
        })

    except Exception as e:
        logger.error(f"Errore participants list API: {str(e)}")
        return jsonify({'participants': [], 'cd_keys': [], 'sf_keys': []})


@admin_bp.route('/workflows/<int:workflow_id>/steps/<int:step_id>/landing-builder')
def landing_builder(workflow_id, step_id):
    """Configuratore landing page"""
    try:
        step = db.query(WorkflowStep).filter_by(id=step_id, workflow_id=workflow_id).first()

        # Fallback: se lo step ID non esiste, cerca per ordine (gli ID cambiano dopo ogni save)
        if not step:
            step = db.query(WorkflowStep).filter_by(workflow_id=workflow_id, order=step_id).first()

        if not step:
            flash('Step non trovato', 'danger')
            return redirect(url_for('admin.workflow_detail', workflow_id=workflow_id))

        return render_template('admin/landing_config.html',
                             step=step,
                             config=step.landing_gjs_data)

    except Exception as e:
        logger.error(f"Errore landing config: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return redirect(url_for('admin.workflow_detail', workflow_id=workflow_id))


@admin_bp.route('/api/collected-data')
def collected_data_api():
    """API endpoint for collected data (AJAX)"""
    import pytz
    local_tz = pytz.timezone('Europe/Rome')

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        LIMIT = 500

        workflows = db.query(Workflow).all()
        wf_names = {w.id: w.name for w in workflows}

        query = db.query(
            Participant.id, Participant.first_name, Participant.last_name,
            Participant.email, Participant.phone, Participant.workflow_id,
            Participant.status, Participant.completed_at, Participant.collected_data,
            Participant.sabaform_data
        ).filter(
            Participant.collected_data.isnot(None)
        )
        if workflow_id:
            query = query.filter(Participant.workflow_id == workflow_id)

        rows = query.order_by(
            Participant.completed_at.is_(None).asc(),
            Participant.completed_at.desc()
        ).limit(LIMIT).all()

        def _fmt_time(dt):
            if dt is None:
                return None
            utc_dt = pytz.utc.localize(dt)
            return utc_dt.astimezone(local_tz).strftime('%d/%m/%Y %H:%M')

        def _strip_file_data(cd):
            """Strip base64 data from file objects, keep metadata only"""
            if not cd or not isinstance(cd, dict):
                return cd or {}
            light = {}
            for k, v in cd.items():
                if isinstance(v, dict) and v.get('filename') and v.get('data'):
                    light[k] = {'filename': v['filename'], 'mime': v.get('mime', ''), 'has_file': True}
                else:
                    light[k] = v
            return light

        participants = []
        for pid, fn, ln, email, phone, wf_id, status, completed_at, cd, sf in rows:
            participants.append({
                'id': pid,
                'first_name': fn or '',
                'last_name': ln or '',
                'email': email or '',
                'phone': phone or '',
                'workflow_id': wf_id,
                'workflow': wf_names.get(wf_id, ''),
                'status': status.value,
                'completed_at': _fmt_time(completed_at),
                'collected_data': _strip_file_data(cd),
            })

        return jsonify(participants)

    except Exception as e:
        logger.error(f"Errore collected data API: {str(e)}")
        return jsonify([])


@admin_bp.route('/api/collected-data/export-all')
def export_all_collected_excel():
    """Export all participants with collected data (same format as completed)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.numbers import FORMAT_TEXT
    from io import BytesIO
    import re
    import pytz

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        if not workflow_id:
            return jsonify({'error': 'workflow_id richiesto'}), 400

        local_tz = pytz.timezone('Europe/Rome')

        # Get landing field order from step config
        steps = db.query(WorkflowStep).filter_by(workflow_id=workflow_id).order_by(WorkflowStep.order).all()
        ordered_fields = []
        seen = set()
        for step in steps:
            for config in [step.landing_page_config, step.landing_gjs_data]:
                if config and isinstance(config, dict):
                    for f in config.get('fields', []):
                        name = f.get('name', '')
                        if name and name not in seen:
                            seen.add(name)
                            ordered_fields.append({'name': name, 'label': f.get('label', name)})

        # Get ALL participants with collected data
        rows = db.query(
            Participant.id, Participant.first_name, Participant.last_name,
            Participant.email, Participant.collected_data
        ).filter(
            Participant.workflow_id == workflow_id,
            Participant.collected_data.isnot(None)
        ).order_by(Participant.last_name, Participant.first_name).all()

        # Latest substate per participant
        SUBSTATE_LABELS = {
            'scheduled': 'Scheduled', 'sent': 'Email Sent', 'delivered': 'Email Sent',
            'opened': 'Email Opened', 'clicked': 'Link Clicked',
            'completed': 'Completed', 'failed': 'Failed',
            'landing_opened': 'Landing Opened', 'form_submitted': 'Form Submitted',
            'survey_submitted': 'Survey Submitted',
        }
        participant_ids = [r[0] for r in rows]
        latest_substate = {}

        if participant_ids:
            for pid_val, evt, created in db.query(
                ActivityLog.participant_id, ActivityLog.event_type, func.max(ActivityLog.created_at)
            ).filter(
                ActivityLog.workflow_id == workflow_id,
                ActivityLog.participant_id.in_(participant_ids),
                ActivityLog.event_type.in_(['landing_opened', 'form_submitted', 'survey_submitted'])
            ).group_by(ActivityLog.participant_id, ActivityLog.event_type).all():
                if pid_val not in latest_substate or (created and created > latest_substate[pid_val][1]):
                    latest_substate[pid_val] = (SUBSTATE_LABELS.get(evt, evt), created)

            step_ids = [s.id for s in steps]
            if step_ids:
                for pid_val, status, ts in db.query(
                    Execution.participant_id, Execution.status,
                    func.max(func.coalesce(Execution.sent_at, Execution.scheduled_at, Execution.created_at))
                ).filter(
                    Execution.step_id.in_(step_ids),
                    Execution.participant_id.in_(participant_ids)
                ).group_by(Execution.participant_id, Execution.status).all():
                    label = SUBSTATE_LABELS.get(status.value, status.value)
                    if pid_val not in latest_substate or (ts and ts > latest_substate[pid_val][1]):
                        latest_substate[pid_val] = (label, ts)

        # Fallback field order
        if not ordered_fields:
            all_keys = []
            keys_seen = set()
            for _, _, _, _, cd in rows:
                if cd and isinstance(cd, dict):
                    for k in cd.keys():
                        if k not in keys_seen:
                            keys_seen.add(k)
                            all_keys.append(k)
            ordered_fields = [{'name': k, 'label': k} for k in all_keys]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Dati Raccolti'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='795548', end_color='795548', fill_type='solid')
        thin_border = Border(bottom=Side(style='thin', color='D7CCC8'))

        headers = ['Nome', 'Cognome', 'Email'] + [f['label'] for f in ordered_fields] + ['Sottostato', 'Data Sottostato']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        date_field_names = {'data_nascita', 'birth_date', 'data_di_nascita', 'date_of_birth'}
        phone_field_names = {'telefono', 'phone', 'cellulare', 'mobile', 'tel'}
        date_re = re.compile(r'^(\d{4})-(\d{2})-(\d{2})')

        def _fmt_ts(ts):
            if ts is None:
                return ''
            utc_dt = pytz.utc.localize(ts)
            return utc_dt.astimezone(local_tz).strftime('%d-%m-%Y %H:%M')

        for row_idx, (pid, fn, ln, email, cd) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=fn or '')
            ws.cell(row=row_idx, column=2, value=ln or '')
            ws.cell(row=row_idx, column=3, value=email or '')

            cd = cd or {}
            for col_offset, field in enumerate(ordered_fields):
                val = cd.get(field['name'], '')
                if field['name'].lower() in date_field_names and isinstance(val, str):
                    m = date_re.match(val)
                    if m:
                        val = f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
                if isinstance(val, dict) and 'filename' in val:
                    val = val.get('filename', '[file]')
                cell = ws.cell(row=row_idx, column=4 + col_offset, value=str(val) if val else '')
                if field['name'].lower() in phone_field_names:
                    cell.number_format = FORMAT_TEXT

            substate_info = latest_substate.get(pid, ('', None))
            ws.cell(row=row_idx, column=len(headers) - 1, value=substate_info[0])
            ws.cell(row=row_idx, column=len(headers), value=_fmt_ts(substate_info[1]))

            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).border = thin_border

        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        wf = db.get(Workflow, workflow_id)
        wf_name = (wf.name if wf else 'export').replace(' ', '_')
        filename = f'dati_raccolti_{wf_name}_{datetime.utcnow().strftime("%Y-%m-%d")}.xlsx'

        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logger.error(f"Errore export all collected: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/collected-data/export-completed')
def export_completed_excel():
    """Export completed participants with landing field order"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        if not workflow_id:
            return jsonify({'error': 'workflow_id richiesto'}), 400

        # Get landing field order from step config
        steps = db.query(WorkflowStep).filter_by(workflow_id=workflow_id).order_by(WorkflowStep.order).all()
        ordered_fields = []
        seen = set()
        for step in steps:
            for config in [step.landing_page_config, step.landing_gjs_data]:
                if config and isinstance(config, dict):
                    for f in config.get('fields', []):
                        name = f.get('name', '')
                        if name and name not in seen:
                            seen.add(name)
                            ordered_fields.append({'name': name, 'label': f.get('label', name)})

        # Get completed participants
        rows = db.query(
            Participant.id, Participant.first_name, Participant.last_name,
            Participant.email, Participant.collected_data
        ).filter(
            Participant.workflow_id == workflow_id,
            Participant.status == ParticipantStatus.COMPLETED,
            Participant.collected_data.isnot(None)
        ).order_by(Participant.last_name, Participant.first_name).all()

        import pytz
        local_tz = pytz.timezone('Europe/Rome')

        # Get latest substate per participant from activity_log + executions
        SUBSTATE_LABELS = {
            'scheduled': 'Scheduled', 'sent': 'Email Sent', 'delivered': 'Email Sent',
            'opened': 'Email Opened', 'clicked': 'Link Clicked',
            'completed': 'Completed', 'failed': 'Failed',
            'landing_opened': 'Landing Opened', 'form_submitted': 'Form Submitted',
            'survey_submitted': 'Survey Submitted',
        }

        participant_ids = [r[0] for r in rows]
        latest_substate = {}  # pid → (label, timestamp)

        if participant_ids:
            # From activity_log
            for pid_val, evt, created in db.query(
                ActivityLog.participant_id, ActivityLog.event_type, func.max(ActivityLog.created_at)
            ).filter(
                ActivityLog.workflow_id == workflow_id,
                ActivityLog.participant_id.in_(participant_ids),
                ActivityLog.event_type.in_(['landing_opened', 'form_submitted', 'survey_submitted'])
            ).group_by(ActivityLog.participant_id, ActivityLog.event_type).all():
                if pid_val not in latest_substate or (created and created > latest_substate[pid_val][1]):
                    latest_substate[pid_val] = (SUBSTATE_LABELS.get(evt, evt), created)

            # From executions (email states)
            step_ids = [s.id for s in steps]
            if step_ids:
                for pid_val, status, ts in db.query(
                    Execution.participant_id, Execution.status,
                    func.max(func.coalesce(Execution.sent_at, Execution.scheduled_at, Execution.created_at))
                ).filter(
                    Execution.step_id.in_(step_ids),
                    Execution.participant_id.in_(participant_ids)
                ).group_by(Execution.participant_id, Execution.status).all():
                    label = SUBSTATE_LABELS.get(status.value, status.value)
                    if pid_val not in latest_substate or (ts and ts > latest_substate[pid_val][1]):
                        latest_substate[pid_val] = (label, ts)

        # If no field order from config, derive from collected_data keys
        if not ordered_fields:
            all_keys = []
            keys_seen = set()
            for _, _, _, _, cd in rows:
                if cd and isinstance(cd, dict):
                    for k in cd.keys():
                        if k not in keys_seen:
                            keys_seen.add(k)
                            all_keys.append(k)
            ordered_fields = [{'name': k, 'label': k} for k in all_keys]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Completati'

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='795548', end_color='795548', fill_type='solid')
        thin_border = Border(bottom=Side(style='thin', color='D7CCC8'))
        from openpyxl.styles.numbers import FORMAT_TEXT

        # Header: Nome, Cognome, Email + landing fields + Sottostato + Data Sottostato
        headers = ['Nome', 'Cognome', 'Email'] + [f['label'] for f in ordered_fields] + ['Sottostato', 'Data Sottostato']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        import re
        date_field_names = {'data_nascita', 'birth_date', 'data_di_nascita', 'date_of_birth'}
        phone_field_names = {'telefono', 'phone', 'cellulare', 'mobile', 'tel'}
        date_re = re.compile(r'^(\d{4})-(\d{2})-(\d{2})')

        def _fmt_ts(ts):
            if ts is None:
                return ''
            utc_dt = pytz.utc.localize(ts)
            return utc_dt.astimezone(local_tz).strftime('%d-%m-%Y %H:%M')

        for row_idx, (pid, fn, ln, email, cd) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=fn or '')
            ws.cell(row=row_idx, column=2, value=ln or '')
            ws.cell(row=row_idx, column=3, value=email or '')

            cd = cd or {}
            for col_offset, field in enumerate(ordered_fields):
                val = cd.get(field['name'], '')
                # Format dates dd-mm-yyyy
                if field['name'].lower() in date_field_names and isinstance(val, str):
                    m = date_re.match(val)
                    if m:
                        val = f'{m.group(3)}-{m.group(2)}-{m.group(1)}'
                # File uploads: show filename only
                if isinstance(val, dict) and 'filename' in val:
                    val = val.get('filename', '[file]')
                cell = ws.cell(row=row_idx, column=4 + col_offset, value=str(val) if val else '')
                # Phone fields: format as text so Excel doesn't strip leading zeros
                if field['name'].lower() in phone_field_names:
                    cell.number_format = FORMAT_TEXT

            # Sottostato (penultima colonna)
            substate_info = latest_substate.get(pid, ('', None))
            ws.cell(row=row_idx, column=len(headers) - 1, value=substate_info[0])
            # Data Sottostato (ultima colonna)
            ws.cell(row=row_idx, column=len(headers), value=_fmt_ts(substate_info[1]))

            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).border = thin_border

        # Auto-width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        wf = db.get(Workflow, workflow_id)
        wf_name = (wf.name if wf else 'export').replace(' ', '_')
        filename = f'completati_{wf_name}_{datetime.utcnow().strftime("%Y-%m-%d")}.xlsx'

        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logger.error(f"Errore export completed: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/collected-data')
def collected_data():
    """Vista dati raccolti da tutti i partecipanti"""
    try:
        workflow_id = request.args.get('workflow_id', type=int)

        workflows = db.query(Workflow).all()

        # Default to most recently updated active workflow
        if not workflow_id and workflows:
            active = [w for w in workflows if w.status.value == 'active']
            if active:
                workflow_id = max(active, key=lambda w: w.updated_at or w.created_at).id
            else:
                workflow_id = max(workflows, key=lambda w: w.updated_at or w.created_at).id

        return render_template('admin/collected_data.html',
                             workflows=workflows,
                             current_workflow_id=workflow_id)

    except Exception as e:
        logger.error(f"Errore collected data: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return render_template('admin/collected_data.html',
                             workflows=[],
                             current_workflow_id=None)


@admin_bp.route('/api/executions/timeline')
def executions_timeline_api():
    """API endpoint for timeline data (AJAX)"""
    import pytz
    local_tz = pytz.timezone('Europe/Rome')

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        TIMELINE_LIMIT = 500

        wf_names = {w.id: w.name for w in db.query(Workflow.id, Workflow.name).all()}

        # Query executions con JOIN (niente mappe in memoria)
        from sqlalchemy.sql import func, case, literal_column
        from sqlalchemy import literal

        exec_query = db.query(
            Execution.id,
            Execution.participant_id,
            Execution.status,
            Execution.scheduled_at,
            Execution.sent_at,
            Execution.error_message,
            Participant.first_name,
            Participant.last_name,
            Participant.email.label('p_email'),
            Participant.workflow_id,
            WorkflowStep.name.label('step_name'),
            WorkflowStep.type.label('step_type'),
        ).join(Participant, Execution.participant_id == Participant.id)\
         .join(WorkflowStep, Execution.step_id == WorkflowStep.id)

        if workflow_id:
            exec_query = exec_query.filter(Participant.workflow_id == workflow_id)
        executions = exec_query.order_by(Execution.scheduled_at.desc()).limit(TIMELINE_LIMIT).all()

        # Query activities con JOIN
        activity_query = db.query(
            ActivityLog.id,
            ActivityLog.participant_id,
            ActivityLog.workflow_id,
            ActivityLog.event_type,
            ActivityLog.description,
            ActivityLog.created_at,
            Participant.first_name,
            Participant.last_name,
            Participant.email.label('p_email'),
            WorkflowStep.name.label('step_name'),
            WorkflowStep.type.label('step_type'),
        ).outerjoin(Participant, ActivityLog.participant_id == Participant.id)\
         .outerjoin(WorkflowStep, ActivityLog.step_id == WorkflowStep.id)

        if workflow_id:
            activity_query = activity_query.filter(ActivityLog.workflow_id == workflow_id)
        activities = activity_query.order_by(ActivityLog.created_at.desc()).limit(TIMELINE_LIMIT).all()

        def _fmt_time(dt):
            if dt is None:
                return ''
            utc_dt = pytz.utc.localize(dt)
            return utc_dt.astimezone(local_tz).strftime('%d/%m/%Y %H:%M')

        def _pname(fn, ln, email):
            name = ' '.join(p for p in [fn or '', ln or ''] if p).strip()
            return name or email or '—'

        timeline = []
        for row in executions:
            display_time = row.sent_at or row.scheduled_at
            stype = row.step_type.value if row.step_type else ''
            sname = row.step_name or '—'
            timeline.append({
                'entry_type': 'execution',
                'entry_id': row.id,
                'participant_id': row.participant_id,
                'time': display_time.isoformat() if display_time else '',
                'time_local': _fmt_time(display_time),
                'event_type': 'email_failed' if row.status.value == 'failed' else 'email_' + row.status.value,
                'description': f'{sname}: {row.status.value}',
                'participant': _pname(row.first_name, row.last_name, row.p_email),
                'step': sname,
                'step_type': stype,
                'workflow': wf_names.get(row.workflow_id, ''),
                'error': row.error_message,
            })
        for row in activities:
            sname = row.step_name or '—'
            stype = row.step_type.value if row.step_type else ''
            timeline.append({
                'entry_type': 'activity',
                'entry_id': row.id,
                'participant_id': row.participant_id,
                'time': row.created_at.isoformat() if row.created_at else '',
                'time_local': _fmt_time(row.created_at),
                'event_type': row.event_type,
                'description': row.description,
                'participant': _pname(row.first_name, row.last_name, row.p_email),
                'step': sname,
                'step_type': stype,
                'workflow': wf_names.get(row.workflow_id, ''),
                'error': None,
            })

        timeline.sort(key=lambda x: x['time'], reverse=True)
        timeline = timeline[:TIMELINE_LIMIT]
        return jsonify(timeline)

    except Exception as e:
        logger.error(f"Errore timeline API: {str(e)}")
        return jsonify([])


@admin_bp.route('/api/participant/<int:pid>/full-data')
def participant_full_data(pid):
    """Fetch full collected_data and sabaform_data for a single participant (on-demand)"""
    try:
        row = db.query(
            Participant.collected_data, Participant.sabaform_data
        ).filter_by(id=pid).first()
        if not row:
            return jsonify({'error': 'Not found'}), 404
        return jsonify({
            'collected_data': row.collected_data or {},
            'sabaform_data': row.sabaform_data or {},
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/timeline-entry-details')
def timeline_entry_details():
    """Fetch details for a single timeline entry (on-demand)"""
    entry_type = request.args.get('type')  # 'execution' or 'activity'
    entry_id = request.args.get('id', type=int)
    if not entry_type or not entry_id:
        return jsonify({'error': 'Missing params'}), 400
    try:
        if entry_type == 'execution':
            row = db.query(Execution.result_data).filter_by(id=entry_id).first()
            return jsonify({'details': row.result_data if row else None})
        else:
            row = db.query(ActivityLog.details).filter_by(id=entry_id).first()
            return jsonify({'details': row.details if row else None})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/executions')
def executions_monitor():
    """Monitor esecuzioni"""
    try:
        workflow_id = request.args.get('workflow_id', type=int)

        # Load workflows with steps in one query
        workflows = db.query(Workflow).options(selectinload(Workflow.steps)).all()

        # Default to most recently updated workflow if none selected
        if not workflow_id and workflows:
            active = [w for w in workflows if w.status.value == 'active']
            if active:
                workflow_id = max(active, key=lambda w: w.updated_at or w.created_at).id
            else:
                workflow_id = max(workflows, key=lambda w: w.updated_at or w.created_at).id

        # Step-based flow data — BATCH QUERIES (avoid N+1)
        flow_data = {}
        target_workflows = [w for w in workflows if w.id == workflow_id] if workflow_id else workflows
        target_wf_ids = [w.id for w in target_workflows]
        if not target_wf_ids:
            target_wf_ids = [0]  # no-op filter

        # Batch 1: participant counts per workflow+status (1 query)
        participant_counts = {}
        for wf_id, status, cnt in db.query(
            Participant.workflow_id, Participant.status, func.count(Participant.id)
        ).filter(Participant.workflow_id.in_(target_wf_ids)).group_by(
            Participant.workflow_id, Participant.status
        ).all():
            participant_counts.setdefault(wf_id, {})[status.value] = cnt

        # Batch 2: execution counts per step+status (1 query)
        all_step_ids = []
        step_map = {}  # step_id → step object
        wf_steps = {}  # wf_id → [step, ...]
        for wf in target_workflows:
            sorted_steps = sorted(wf.steps, key=lambda s: s.order)
            wf_steps[wf.id] = sorted_steps
            for s in sorted_steps:
                all_step_ids.append(s.id)
                step_map[s.id] = s

        exec_batch = {}  # step_id → {status: count}
        if all_step_ids:
            for step_id, status, cnt in db.query(
                Execution.step_id, Execution.status, func.count(Execution.id)
            ).filter(Execution.step_id.in_(all_step_ids)).group_by(
                Execution.step_id, Execution.status
            ).all():
                exec_batch.setdefault(step_id, {})[status.value] = cnt

        # Batch 3: activity counts per step+event (distinct participants) (1 query)
        activity_by_step = {}  # step_id → {event_type: count}
        if all_step_ids:
            for step_id, evt, cnt in db.query(
                ActivityLog.step_id, ActivityLog.event_type,
                func.count(func.distinct(ActivityLog.participant_id))
            ).filter(
                ActivityLog.step_id.in_(all_step_ids),
                ActivityLog.participant_id.isnot(None),
                ActivityLog.event_type.in_(['landing_opened', 'form_submitted', 'survey_submitted'])
            ).group_by(ActivityLog.step_id, ActivityLog.event_type).all():
                activity_by_step.setdefault(step_id, {})[evt] = cnt

        # Batch 4: fallback activity counts per workflow (no step_id, legacy) (1 query)
        activity_by_wf_legacy = {}  # wf_id → {event_type: count}
        for wf_id, evt, cnt in db.query(
            ActivityLog.workflow_id, ActivityLog.event_type,
            func.count(func.distinct(ActivityLog.participant_id))
        ).filter(
            ActivityLog.workflow_id.in_(target_wf_ids),
            ActivityLog.participant_id.isnot(None),
            ActivityLog.step_id.is_(None),
            ActivityLog.event_type.in_(['landing_opened', 'form_submitted', 'survey_submitted'])
        ).group_by(ActivityLog.workflow_id, ActivityLog.event_type).all():
            activity_by_wf_legacy.setdefault(wf_id, {})[evt] = cnt

        # Batch 5: current_step counts per workflow+step (1 query)
        current_at_batch = {}  # (wf_id, step_id) → count
        for wf_id, step_id, cnt in db.query(
            Participant.workflow_id, Participant.current_step_id, func.count(Participant.id)
        ).filter(
            Participant.workflow_id.in_(target_wf_ids),
            Participant.current_step_id.isnot(None)
        ).group_by(Participant.workflow_id, Participant.current_step_id).all():
            current_at_batch[(wf_id, step_id)] = cnt

        # Build flow_data from batched results (no more queries in loops)
        for wf in target_workflows:
            wf_pcounts = participant_counts.get(wf.id, {})
            total = sum(wf_pcounts.values())
            if total == 0:
                continue

            steps = wf_steps.get(wf.id, [])
            # Pre-compute first landing/survey step for fallback logic
            first_landing_id = next((s.id for s in steps if s.landing_html or s.landing_gjs_data or s.landing_page_config), None)
            first_survey_id = next((s.id for s in steps if s.type.value == 'survey'), None)

            steps_flow = []
            for step in steps:
                exec_map = exec_batch.get(step.id, {})
                has_landing = bool(step.landing_html or step.landing_gjs_data or step.landing_page_config)

                # Activity map: step-specific + legacy (step_id=NULL) merged for first landing/survey
                activity_map = dict(activity_by_step.get(step.id, {}))
                is_first_landing = has_landing and first_landing_id == step.id
                is_first_survey = step.type.value == 'survey' and first_survey_id == step.id
                if (is_first_landing or is_first_survey) and activity_by_wf_legacy.get(wf.id):
                    relevant = []
                    if is_first_landing:
                        relevant.extend(['landing_opened', 'form_submitted'])
                    if is_first_survey:
                        relevant.append('survey_submitted')
                    legacy = activity_by_wf_legacy[wf.id]
                    has_legacy = any(k in legacy for k in relevant)
                    if has_legacy:
                        # Query distinct count across step_id=X and step_id=NULL to avoid double counting
                        for evt, cnt in db.query(
                            ActivityLog.event_type,
                            func.count(func.distinct(ActivityLog.participant_id))
                        ).filter(
                            ActivityLog.workflow_id == wf.id,
                            ActivityLog.participant_id.isnot(None),
                            ActivityLog.event_type.in_(relevant),
                            (ActivityLog.step_id == step.id) | (ActivityLog.step_id.is_(None))
                        ).group_by(ActivityLog.event_type).all():
                            activity_map[evt] = cnt

                # Build substates
                substates = []
                sent = exec_map.get('sent', 0) + exec_map.get('delivered', 0)
                scheduled = exec_map.get('scheduled', 0)
                opened = exec_map.get('opened', 0)
                clicked = exec_map.get('clicked', 0)
                completed = exec_map.get('completed', 0)
                failed = exec_map.get('failed', 0)
                skipped = exec_map.get('skipped', 0)
                total_entered = sum(exec_map.values()) - skipped

                if scheduled:
                    substates.append({'key': 'scheduled', 'label': 'Scheduled', 'count': scheduled, 'icon': 'bi-clock', 'color': '#ff9800'})
                if sent:
                    substates.append({'key': 'sent', 'label': 'Email Sent', 'count': sent, 'icon': 'bi-envelope-check', 'color': '#2196f3'})
                if opened:
                    substates.append({'key': 'opened', 'label': 'Email Opened', 'count': opened, 'icon': 'bi-envelope-open', 'color': '#00bcd4'})
                if clicked:
                    substates.append({'key': 'clicked', 'label': 'Link Clicked', 'count': clicked, 'icon': 'bi-cursor', 'color': '#9c27b0'})
                if activity_map.get('landing_opened', 0):
                    substates.append({'key': 'landing_opened', 'label': 'Landing Opened', 'count': activity_map['landing_opened'], 'icon': 'bi-box-arrow-up-right', 'color': '#ff9800'})
                if activity_map.get('form_submitted', 0):
                    substates.append({'key': 'form_submitted', 'label': 'Form Submitted', 'count': activity_map['form_submitted'], 'icon': 'bi-check2-square', 'color': '#4caf50'})
                if activity_map.get('survey_submitted', 0):
                    substates.append({'key': 'survey_submitted', 'label': 'Survey Submitted', 'count': activity_map['survey_submitted'], 'icon': 'bi-ui-checks', 'color': '#00bcd4'})
                if completed:
                    substates.append({'key': 'completed', 'label': 'Completed', 'count': completed, 'icon': 'bi-check-all', 'color': '#4caf50'})
                if failed:
                    substates.append({'key': 'failed', 'label': 'Failed', 'count': failed, 'icon': 'bi-x-circle', 'color': '#f44336'})

                steps_flow.append({
                    'id': step.id,
                    'order': step.order,
                    'name': step.name,
                    'type': step.type.value,
                    'entered': total_entered,
                    'current_at': current_at_batch.get((wf.id, step.id), 0),
                    'substates': substates,
                })

            flow_data[wf.id] = {
                'name': wf.name,
                'total': total,
                'steps': steps_flow,
                'status_counts': wf_pcounts,
            }

        return render_template('admin/executions_monitor.html',
                             workflows=workflows,
                             current_workflow_id=workflow_id,
                             flow_data=flow_data)

    except Exception as e:
        logger.error(f"Errore monitor esecuzioni: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return render_template('admin/executions_monitor.html',
                             workflows=[],
                             current_workflow_id=None,
                             flow_data={})


@admin_bp.route('/activity-log')
@superuser_required
def activity_log():
    """User audit log — superuser only"""
    from app.translations import get_translations
    from app.models import UserAuditLog
    t = get_translations()
    try:
        action_filter = request.args.get('action', '')
        entity_filter = request.args.get('entity', '')
        user_filter = request.args.get('user', '').strip()
        days = request.args.get('days', 30, type=int)

        from datetime import timedelta
        cutoff = datetime.utcnow() - timedelta(days=days)

        query = db.query(UserAuditLog).filter(UserAuditLog.timestamp >= cutoff)

        if action_filter:
            query = query.filter(UserAuditLog.action == action_filter)
        if entity_filter:
            query = query.filter(UserAuditLog.entity == entity_filter)
        if user_filter:
            query = query.filter(UserAuditLog.user_email.ilike(f'%{user_filter}%'))

        logs = query.order_by(UserAuditLog.timestamp.desc()).limit(500).all()

        all_actions = [r[0] for r in db.query(UserAuditLog.action).distinct().order_by(UserAuditLog.action).all()]
        all_entities = [r[0] for r in db.query(UserAuditLog.entity).distinct().order_by(UserAuditLog.entity).all() if r[0]]

        return render_template('admin/activity_log.html',
                             logs=logs,
                             all_actions=all_actions,
                             all_entities=all_entities,
                             action_filter=action_filter,
                             entity_filter=entity_filter,
                             user_filter=user_filter,
                             days_filter=days,
                             t=t)
    except Exception as e:
        logger.error(f"Errore activity log: {str(e)}")
        flash(f'Errore: {str(e)}', 'danger')
        return render_template('admin/activity_log.html',
                             logs=[], all_actions=[], all_entities=[],
                             action_filter='', entity_filter='',
                             user_filter='', days_filter=30, t=t)


ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg'}

MIME_TYPES = {
    'png': 'image/png',
    'jpg': 'image/jpeg',
    'jpeg': 'image/jpeg',
    'gif': 'image/gif',
    'webp': 'image/webp',
    'svg': 'image/svg+xml',
}


@admin_bp.route('/upload-image', methods=['POST'])
def upload_image():
    """Upload immagine per landing page (logo o sfondo) — salva in DB"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file selezionato'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nessun file selezionato'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return jsonify({'error': f'Formato non consentito. Usa: {", ".join(ALLOWED_IMAGE_EXTENSIONS)}'}), 400

    filename = f"{uuid.uuid4().hex}.{ext}"
    image = UploadedImage(
        filename=filename,
        mime_type=MIME_TYPES.get(ext, 'application/octet-stream'),
        data=file.read()
    )
    db.add(image)
    db.commit()

    url = url_for('admin.serve_image', image_id=image.id, _external=False)
    return jsonify({'url': url}), 200


@admin_bp.route('/images/<int:image_id>')
def serve_image(image_id):
    """Serve immagine dal DB"""
    image = db.query(UploadedImage).get(image_id)
    if not image:
        return 'Not found', 404
    return Response(
        image.data,
        mimetype=image.mime_type,
        headers={'Cache-Control': 'public, max-age=31536000'}
    )


ALLOWED_ATTACHMENT_EXTENSIONS = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'txt', 'png', 'jpg', 'jpeg', 'gif', 'zip'}
MAX_ATTACHMENT_SIZE = int(3.9 * 1024 * 1024)  # 3.9 MB (Graph API limit)


@admin_bp.route('/api/attachments', methods=['POST'])
def upload_attachment():
    """Upload allegato per step email"""
    if 'file' not in request.files:
        return jsonify({'error': 'Nessun file selezionato'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nessun file selezionato'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_ATTACHMENT_EXTENSIONS:
        return jsonify({'error': f'Formato non consentito. Usa: {", ".join(sorted(ALLOWED_ATTACHMENT_EXTENSIONS))}'}), 400

    data = file.read()
    if len(data) > MAX_ATTACHMENT_SIZE:
        return jsonify({'error': 'File troppo grande (max 10 MB)'}), 400

    attachment = Attachment(
        filename=file.filename,
        mime_type=file.content_type or 'application/octet-stream',
        size=len(data),
        data=data
    )
    db.add(attachment)
    db.commit()

    return jsonify({
        'id': attachment.id,
        'filename': attachment.filename,
        'size': attachment.size,
        'mime_type': attachment.mime_type
    }), 201


@admin_bp.route('/api/attachments/<int:attachment_id>', methods=['DELETE'])
def delete_attachment(attachment_id):
    """Elimina allegato"""
    attachment = db.get(Attachment, attachment_id)
    if not attachment:
        return jsonify({'error': 'Not found'}), 404

    db.delete(attachment)
    db.commit()
    return jsonify({'success': True}), 200


@admin_bp.route('/api/attachments/info', methods=['POST'])
def get_attachments_info():
    """Restituisce metadati degli allegati dato un array di IDs"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify([]), 200

    attachments = db.query(Attachment).filter(Attachment.id.in_(ids)).all()
    return jsonify([{
        'id': a.id,
        'filename': a.filename,
        'size': a.size,
        'mime_type': a.mime_type
    } for a in attachments]), 200


@admin_bp.route('/api/step/<int:step_id>/participants')
def step_participants(step_id):
    """Return participants for a given step + substate"""
    try:
        substate = request.args.get('substate', '')
        step = db.get(WorkflowStep, step_id)
        if not step:
            return jsonify({'error': 'Step not found'}), 404

        # Activity-based substates
        ACTIVITY_SUBSTATES = {'landing_opened', 'form_submitted', 'survey_submitted'}
        # Execution-based substates
        EXEC_SUBSTATES = {'scheduled', 'sent', 'delivered', 'opened', 'clicked', 'completed', 'failed', 'skipped'}

        participants = []
        if substate == 'current_at':
            # Participants whose current_step_id is this step
            parts = db.query(Participant).filter(
                Participant.workflow_id == step.workflow_id,
                Participant.current_step_id == step_id
            ).all()
            participants = [{'id': p.id, 'name': p.full_name or p.email or f'#{p.id}', 'email': p.email or '', 'status': p.status.value} for p in parts]
        elif substate in ACTIVITY_SUBSTATES:
            # Distinct participants from ActivityLog
            rows = db.query(ActivityLog.participant_id).filter(
                ActivityLog.workflow_id == step.workflow_id,
                ActivityLog.event_type == substate,
                ActivityLog.participant_id.isnot(None)
            )
            if substate != 'landing_opened':
                # For non-landing events, also filter by step if available
                rows = rows.filter(
                    (ActivityLog.step_id == step_id) | (ActivityLog.step_id.is_(None))
                )
            pids = [r[0] for r in rows.distinct().all()]
            if pids:
                parts = db.query(Participant).filter(Participant.id.in_(pids)).all()
                participants = [{'id': p.id, 'name': p.full_name or p.email or f'#{p.id}', 'email': p.email or '', 'status': p.status.value} for p in parts]
        elif substate in EXEC_SUBSTATES:
            # Map 'sent' to include 'delivered' too
            statuses = [substate]
            if substate == 'sent':
                statuses.append('delivered')
            pids = [r[0] for r in db.query(Execution.participant_id).filter(
                Execution.step_id == step_id,
                Execution.status.in_(statuses)
            ).distinct().all()]
            if pids:
                parts = db.query(Participant).filter(Participant.id.in_(pids)).all()
                participants = [{'id': p.id, 'name': p.full_name or p.email or f'#{p.id}', 'email': p.email or '', 'status': p.status.value} for p in parts]

        return jsonify({'participants': participants, 'step': step.name, 'substate': substate}), 200
    except Exception as e:
        logger.error(f"Errore step participants: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/workflow/<int:wf_id>/participants')
def workflow_participants_by_status(wf_id):
    """Return participants for a workflow filtered by status"""
    try:
        status = request.args.get('status', '')
        query = db.query(Participant).filter(Participant.workflow_id == wf_id)
        if status:
            query = query.filter(Participant.status == status)
        parts = query.all()
        participants = [{'id': p.id, 'name': p.full_name or p.email or f'#{p.id}', 'email': p.email or '', 'status': p.status.value} for p in parts]
        return jsonify({'participants': participants}), 200
    except Exception as e:
        logger.error(f"Errore workflow participants: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/status-flow/export')
def export_status_flow():
    """Export status flow to Excel with participant lists"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    try:
        workflow_id = request.args.get('workflow_id', type=int)
        target_workflows = db.query(Workflow).options(
            selectinload(Workflow.steps)
        ).all()
        if workflow_id:
            target_workflows = [w for w in target_workflows if w.id == workflow_id]

        wb = Workbook()

        # ── Styles ──
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='795548', end_color='795548', fill_type='solid')
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color='EFEBE9', end_color='EFEBE9', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='D7CCC8')
        )

        FUNNEL_ORDER = ['scheduled', 'sent', 'opened', 'clicked', 'landing_opened', 'form_submitted', 'survey_submitted', 'completed']
        EXEC_SUBSTATES = {'scheduled', 'sent', 'delivered', 'opened', 'clicked', 'completed', 'failed', 'skipped'}
        ACTIVITY_SUBSTATES = {'landing_opened', 'form_submitted', 'survey_submitted'}
        SUBSTATE_LABELS = {
            'scheduled': 'Scheduled', 'sent': 'Email Sent', 'delivered': 'Email Sent',
            'opened': 'Email Opened', 'clicked': 'Link Clicked',
            'completed': 'Completed', 'failed': 'Failed',
            'landing_opened': 'Landing Opened', 'form_submitted': 'Form Submitted',
            'survey_submitted': 'Survey Submitted',
        }

        import pytz
        _local_tz = pytz.timezone('Europe/Rome')

        def _fmt_ts(ts):
            if ts is None:
                return ''
            utc_dt = pytz.utc.localize(ts)
            return utc_dt.astimezone(_local_tz).strftime('%d-%m-%Y %H:%M')

        def _raw_ts(ts):
            """Return localized datetime for sorting (None-safe)"""
            if ts is None:
                return datetime.min
            return pytz.utc.localize(ts).astimezone(_local_tz)

        def _get_participants_for_substate(step, substate, wf):
            """Return list of (name, email, datetime_formatted, raw_ts, status) for a step+substate"""
            if substate in ACTIVITY_SUBSTATES:
                rows = db.query(ActivityLog.participant_id, ActivityLog.created_at).filter(
                    ActivityLog.workflow_id == wf.id,
                    ActivityLog.event_type == substate,
                    ActivityLog.participant_id.isnot(None)
                )
                if substate != 'landing_opened':
                    rows = rows.filter(
                        (ActivityLog.step_id == step.id) | (ActivityLog.step_id.is_(None))
                    )
                pid_ts = {}
                for r in rows.all():
                    if r[0] not in pid_ts or (r[1] and r[1] > pid_ts[r[0]]):
                        pid_ts[r[0]] = r[1]
                pids = list(pid_ts.keys())
            elif substate in EXEC_SUBSTATES:
                statuses = [substate]
                if substate == 'sent':
                    statuses.append('delivered')
                exec_rows = db.query(Execution.participant_id, Execution.sent_at, Execution.scheduled_at, Execution.created_at).filter(
                    Execution.step_id == step.id,
                    Execution.status.in_(statuses)
                ).all()
                pid_ts = {}
                for r in exec_rows:
                    ts = r[1] or r[2] or r[3]
                    if r[0] not in pid_ts or (ts and ts > pid_ts[r[0]]):
                        pid_ts[r[0]] = ts
                pids = list(pid_ts.keys())
            else:
                return []
            if not pids:
                return []
            parts = db.query(Participant).filter(Participant.id.in_(pids)).all()
            result = [(p.full_name or p.email or f'#{p.id}', p.email or '', _fmt_ts(pid_ts.get(p.id)), _raw_ts(pid_ts.get(p.id)), p.status.value) for p in parts]
            result.sort(key=lambda x: x[3])  # sort by timestamp ascending
            return result

        def _apply_header(ws, headers):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

        # ═══════════════════════════════════════
        # Single sheet: Nello Stato (funnel with participants)
        # ═══════════════════════════════════════
        ws = wb.active
        ws.title = 'Nello Stato'
        _apply_header(ws, ['Workflow', 'Sottostato', 'Entrati', 'Nello Stato', 'Partecipante', 'Email', 'Data e Ora', 'Stato'])

        row = 2
        for wf in target_workflows:
            if not wf.steps:
                continue

            # Build funnel: collect all substates across steps
            funnel = []
            for step in sorted(wf.steps, key=lambda s: s.order):
                exec_counts = db.query(
                    Execution.status, func.count(Execution.id)
                ).filter(Execution.step_id == step.id).group_by(Execution.status).all()
                exec_map = {s.value: c for s, c in exec_counts}

                has_landing = bool(step.landing_html or step.landing_gjs_data or step.landing_page_config)
                relevant_events = []
                if has_landing:
                    relevant_events.extend(['landing_opened', 'form_submitted'])
                if step.type.value == 'survey':
                    relevant_events.append('survey_submitted')

                sent = exec_map.get('sent', 0) + exec_map.get('delivered', 0)
                if exec_map.get('scheduled', 0):
                    funnel.append({'key': 'scheduled', 'count': exec_map['scheduled'], 'step': step})
                if sent:
                    funnel.append({'key': 'sent', 'count': sent, 'step': step})
                if exec_map.get('opened', 0):
                    funnel.append({'key': 'opened', 'count': exec_map['opened'], 'step': step})
                if exec_map.get('clicked', 0):
                    funnel.append({'key': 'clicked', 'count': exec_map['clicked'], 'step': step})

                for evt in relevant_events:
                    cnt = db.query(func.count(func.distinct(ActivityLog.participant_id))).filter(
                        ActivityLog.workflow_id == wf.id,
                        ActivityLog.event_type == evt,
                        ActivityLog.participant_id.isnot(None),
                        (ActivityLog.step_id == step.id) | (ActivityLog.step_id.is_(None))
                    ).scalar() or 0
                    if cnt:
                        funnel.append({'key': evt, 'count': cnt, 'step': step})

                if exec_map.get('completed', 0):
                    funnel.append({'key': 'completed', 'count': exec_map['completed'], 'step': step})

            # Sort by funnel order
            funnel.sort(key=lambda f: FUNNEL_ORDER.index(f['key']) if f['key'] in FUNNEL_ORDER else 99)

            if not funnel:
                continue

            # Compute "nello stato"
            for i in range(len(funnel)):
                next_count = funnel[i + 1]['count'] if i < len(funnel) - 1 else 0
                funnel[i]['in_state'] = max(0, funnel[i]['count'] - next_count)

            # Get participants "in state" = participants in this substate but NOT in the next
            for i, f in enumerate(funnel):
                all_pids = set()
                next_pids = set()

                # PIDs for this substate (with timestamps)
                pid_ts = {}
                if f['key'] in ACTIVITY_SUBSTATES:
                    rows = db.query(ActivityLog.participant_id, ActivityLog.created_at).filter(
                        ActivityLog.workflow_id == wf.id,
                        ActivityLog.event_type == f['key'],
                        ActivityLog.participant_id.isnot(None)
                    ).all()
                    for r in rows:
                        if r[0] not in pid_ts or (r[1] and r[1] > pid_ts[r[0]]):
                            pid_ts[r[0]] = r[1]
                    all_pids = set(pid_ts.keys())
                elif f['key'] in EXEC_SUBSTATES:
                    statuses = [f['key']]
                    if f['key'] == 'sent':
                        statuses.append('delivered')
                    exec_rows = db.query(Execution.participant_id, Execution.sent_at, Execution.scheduled_at, Execution.created_at).filter(
                        Execution.step_id == f['step'].id,
                        Execution.status.in_(statuses)
                    ).all()
                    for r in exec_rows:
                        ts = r[1] or r[2] or r[3]
                        if r[0] not in pid_ts or (ts and ts > pid_ts[r[0]]):
                            pid_ts[r[0]] = ts
                    all_pids = set(pid_ts.keys())

                # PIDs for next substate (to subtract)
                if i < len(funnel) - 1:
                    nf = funnel[i + 1]
                    if nf['key'] in ACTIVITY_SUBSTATES:
                        rows = db.query(ActivityLog.participant_id).filter(
                            ActivityLog.workflow_id == wf.id,
                            ActivityLog.event_type == nf['key'],
                            ActivityLog.participant_id.isnot(None)
                        ).distinct().all()
                        next_pids = {r[0] for r in rows}
                    elif nf['key'] in EXEC_SUBSTATES:
                        statuses = [nf['key']]
                        if nf['key'] == 'sent':
                            statuses.append('delivered')
                        rows = db.query(Execution.participant_id).filter(
                            Execution.step_id == nf['step'].id,
                            Execution.status.in_(statuses)
                        ).distinct().all()
                        next_pids = {r[0] for r in rows}

                in_state_pids = all_pids - next_pids
                if in_state_pids:
                    parts = db.query(Participant).filter(Participant.id.in_(in_state_pids)).all()
                    result = [(p.full_name or p.email or f'#{p.id}', p.email or '', _fmt_ts(pid_ts.get(p.id)), _raw_ts(pid_ts.get(p.id)), p.status.value) for p in parts]
                    result.sort(key=lambda x: x[3])  # sort by timestamp ascending
                    f['participants'] = result
                else:
                    f['participants'] = []

            # Write funnel rows
            ws.cell(row=row, column=1, value=wf.name).font = section_font
            ws.cell(row=row, column=1).fill = section_fill
            for c in range(2, 9):
                ws.cell(row=row, column=c).fill = section_fill
            row += 1

            for f in funnel:
                label = SUBSTATE_LABELS.get(f['key'], f['key'])
                if not f['participants']:
                    ws.cell(row=row, column=2, value=label)
                    ws.cell(row=row, column=3, value=f['count'])
                    ws.cell(row=row, column=4, value=f['in_state'])
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).border = thin_border
                    row += 1
                else:
                    for i, (name, email, ts_str, _raw, status) in enumerate(f['participants']):
                        if i == 0:
                            ws.cell(row=row, column=2, value=label)
                            ws.cell(row=row, column=3, value=f['count'])
                            ws.cell(row=row, column=4, value=f['in_state'])
                        ws.cell(row=row, column=5, value=name)
                        ws.cell(row=row, column=6, value=email)
                        ws.cell(row=row, column=7, value=ts_str)
                        ws.cell(row=row, column=8, value=status)
                        for c in range(1, 9):
                            ws.cell(row=row, column=c).border = thin_border
                        row += 1

            row += 1  # blank row between workflows

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        # Write to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f'status_flow_{datetime.utcnow().strftime("%Y-%m-%d")}.xlsx'
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        logger.error(f"Errore export status flow: {str(e)}")
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/participant/<int:participant_id>/timeline')
def participant_timeline(participant_id):
    """Timeline eventi per un partecipante"""
    try:
        participant = db.get(Participant, participant_id)
        if not participant:
            return jsonify({'error': 'Not found'}), 404

        USER_EVENTS = {'form_submitted', 'survey_submitted', 'unsubscribed', 'landing_opened', 'approval_granted', 'approval_rejected'}

        events = []

        # Execution records (ultimi 200, ordinati per data)
        executions = db.query(Execution).options(
            joinedload(Execution.step)
        ).filter_by(participant_id=participant_id)\
         .order_by(Execution.scheduled_at.desc())\
         .limit(200).all()
        for ex in executions:
            step_name = ex.step.name if ex.step else '?'
            step_type = ex.step.type.value if ex.step else '?'
            ts = ex.sent_at or ex.scheduled_at or ex.created_at

            icon = 'envelope'
            if step_type == 'survey':
                icon = 'ui-checks'
            elif step_type == 'goal_check':
                icon = 'trophy'
            elif step_type == 'condition':
                icon = 'shuffle'
            elif step_type == 'wait_until':
                icon = 'calendar-check'

            status_colors = {
                'SCHEDULED': '#6c757d', 'SENT': '#198754', 'FAILED': '#dc3545',
                'SKIPPED': '#ffc107', 'COMPLETED': '#0d6efd'
            }

            events.append({
                'timestamp': ts.isoformat() + 'Z' if ts else None,
                'category': 'system',
                'event_type': ex.status.value if ex.status else 'unknown',
                'description': f'{step_name}',
                'step_type': step_type,
                'icon': icon,
                'color': status_colors.get(ex.status.value if ex.status else '', '#999'),
                'error': ex.error_message,
                'details': ex.result_data
            })

        # ActivityLog records (ultimi 200, ordinati per data)
        activities = db.query(ActivityLog).filter_by(
            participant_id=participant_id
        ).order_by(ActivityLog.created_at.desc()).limit(200).all()
        for a in activities:
            category = 'user' if a.event_type in USER_EVENTS else 'system'

            icon_map = {
                'workflow_started': 'play-circle',
                'form_submitted': 'check2-square',
                'survey_submitted': 'ui-checks',
                'unsubscribed': 'person-x',
                'status_changed': 'arrow-repeat',
                'condition_evaluated': 'shuffle',
                'simulation': 'eye',
                'landing_opened': 'box-arrow-up-right',
                'approval_granted': 'check-circle',
                'approval_rejected': 'x-circle',
                'reconciled': 'arrow-repeat',
            }
            color_map = {
                'workflow_started': '#0d6efd',
                'form_submitted': '#198754',
                'survey_submitted': '#0dcaf0',
                'unsubscribed': '#dc3545',
                'status_changed': '#6c757d',
                'condition_evaluated': '#6c757d',
                'simulation': '#ffc107',
                'landing_opened': '#ff9800',
                'approval_granted': '#198754',
                'approval_rejected': '#dc3545',
                'reconciled': '#fd7e14',
            }

            events.append({
                'timestamp': a.created_at.isoformat() + 'Z' if a.created_at else None,
                'category': category,
                'event_type': a.event_type,
                'description': a.description or a.event_type,
                'icon': icon_map.get(a.event_type, 'info-circle'),
                'color': color_map.get(a.event_type, '#999'),
                'details': a.details
            })

        # Sort by timestamp desc
        events.sort(key=lambda e: e['timestamp'] or '', reverse=True)

        return jsonify({
            'participant': {
                'id': participant.id,
                'name': participant.full_name or participant.email,
                'email': participant.email,
                'status': participant.status.value,
                'enrolled_at': participant.enrolled_at.isoformat() + 'Z' if participant.enrolled_at else None,
                'completed_at': participant.completed_at.isoformat() + 'Z' if participant.completed_at else None,
            },
            'events': events
        })

    except Exception as e:
        logger.error(f"Errore timeline: {str(e)}")
        return jsonify({'error': str(e)}), 500


# =============================================
# USER MANAGEMENT (superuser only)
# =============================================

@admin_bp.route('/users')
@superuser_required
def users_list():
    """Lista utenti"""
    users = db.query(User).order_by(User.created_at.desc()).all()
    workflows = db.query(Workflow).order_by(Workflow.name).all()
    return render_template('admin/users.html', users=users, workflows=workflows)


@admin_bp.route('/users/create', methods=['POST'])
@superuser_required
def user_create():
    """Crea nuovo utente"""
    try:
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        is_superuser = request.form.get('is_superuser') == 'on'

        if not username or not password:
            flash('Username and password required', 'danger')
            return redirect(url_for('admin.users_list'))

        if db.query(User).filter_by(username=username).first():
            flash(f'Username "{username}" already exists', 'danger')
            return redirect(url_for('admin.users_list'))

        user = User(username=username, email=email, is_superuser=is_superuser)
        user.set_password(password)
        db.add(user)
        db.commit()

        from app.services.audit_service import log_user_action
        log_user_action('CREATE', 'User', user.id, f'Created user "{username}"')
        flash(f'User "{username}" created', 'success')
        return redirect(url_for('admin.users_list'))

    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin.users_list'))


@admin_bp.route('/users/<int:user_id>/delete', methods=['POST'])
@superuser_required
def user_delete(user_id):
    """Elimina utente"""
    try:
        from flask import g
        user = db.get(User, user_id)
        if not user:
            flash('User not found', 'danger')
        elif user.id == g.user.id:
            flash('Cannot delete yourself', 'danger')
        else:
            uname = user.username
            db.delete(user)
            db.commit()
            from app.services.audit_service import log_user_action
            log_user_action('DELETE', 'User', user_id, f'Deleted user "{uname}"')
            flash(f'User "{uname}" deleted', 'success')
        return redirect(url_for('admin.users_list'))
    except Exception as e:
        db.rollback()
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('admin.users_list'))


@admin_bp.route('/users/<int:user_id>/toggle-workflow/<int:workflow_id>', methods=['POST'])
@superuser_required
def toggle_user_workflow(user_id, workflow_id):
    """Assign/remove workflow from user"""
    try:
        user = db.get(User, user_id)
        workflow = db.get(Workflow, workflow_id)
        if not user or not workflow:
            return jsonify({'error': 'Not found'}), 404

        if workflow in user.workflows:
            user.workflows.remove(workflow)
            assigned = False
        else:
            user.workflows.append(workflow)
            assigned = True

        db.commit()
        return jsonify({'assigned': assigned}), 200

    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/onedrive/browse')
def onedrive_browse():
    """Browse OneDrive/SharePoint files via Microsoft Graph API."""
    try:
        import requests as http_requests
        from app.services.email_service import EmailService

        storage = request.args.get('storage', 'onedrive')
        folder_path = request.args.get('path', '')
        sharepoint_site = request.args.get('site', '')

        token = EmailService._get_access_token()
        from_email = current_app.config.get('MAIL_FROM_EMAIL', '')
        headers = {"Authorization": f"Bearer {token}"}

        if storage == 'sharepoint':
            if not sharepoint_site:
                return jsonify({'error': 'SharePoint site required'}), 400
            site_url = f"https://graph.microsoft.com/v1.0/sites/{sharepoint_site.strip('/')}"
            site_resp = http_requests.get(site_url, headers=headers, timeout=15)
            if site_resp.status_code != 200:
                return jsonify({'error': 'SharePoint site not found'}), 404
            site_id = site_resp.json()['id']
            drive_base = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive"
        else:
            drive_base = f"https://graph.microsoft.com/v1.0/users/{from_email}/drive"

        if folder_path:
            url = f"{drive_base}/root:/{folder_path.strip('/')}:/children"
        else:
            url = f"{drive_base}/root/children"

        url += "?$select=name,id,folder,file,size,lastModifiedDateTime&$orderby=name"
        resp = http_requests.get(url, headers=headers, timeout=15)

        if resp.status_code != 200:
            return jsonify({'error': f'Graph API error: {resp.status_code}'}), resp.status_code

        items = []
        for item in resp.json().get('value', []):
            is_folder = 'folder' in item
            is_excel = not is_folder and item.get('name', '').lower().endswith(('.xlsx', '.xls'))
            if is_folder or is_excel:
                items.append({
                    'name': item['name'],
                    'id': item['id'],
                    'type': 'folder' if is_folder else 'file',
                    'size': item.get('size', 0),
                    'modified': item.get('lastModifiedDateTime', '')
                })

        return jsonify({'items': items, 'path': folder_path or '/'})

    except Exception as e:
        logger.error(f"OneDrive browse error: {str(e)}")
        return jsonify({'error': str(e)}), 500
